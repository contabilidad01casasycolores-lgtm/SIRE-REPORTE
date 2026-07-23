#!/usr/bin/env python3
"""
Cruza SIRE (TXT con pipes |) vs SAP (Excel) y genera index.html.
Uso: python generar_reporte.py SIRE.txt SAP.xlsx
Tipos considerados: 01, 07, 08, 30, 42 — solo mes actual
"""
import sys, os, json
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("❌ pip install openpyxl"); sys.exit(1)

if len(sys.argv) < 3:
    print("Uso: python generar_reporte.py SIRE.txt SAP.xlsx"); sys.exit(0)

SIRE_FILE = sys.argv[1]
SAP_FILE  = sys.argv[2]

for f in [SIRE_FILE, SAP_FILE]:
    if not os.path.exists(f):
        print(f"❌ No encontrado: {f}"); sys.exit(1)

print(f"📂 SIRE: {SIRE_FILE}")
print(f"📂 SAP:  {SAP_FILE}")

# ── Índices TXT SIRE (separado por |) ────────────────────────────────────────
T_PERIODO=2; T_FECHA=4; T_TIPO=6; T_SERIE=7; T_NRO=9
T_RUC=12; T_RAZON=13; T_BI=14; T_IGV=15; T_NG=20; T_TOTAL=24; T_MONEDA=25
T_ORIG_FECHA=27; T_ORIG_TIPO=28; T_ORIG_SERIE=29; T_ORIG_NRO=31

# ── Índices Excel SAP ─────────────────────────────────────────────────────────
A_TIPO=5; A_SERIE=6; A_NRO=8; A_RUC=11

# ── Reglas de cruce ──────────────────────────────────────────────────────────
# SIRE: 01/07/08/30/42 con serie que empieza con LETRA → Tipo+Serie+N°+RUC
# SIRE: 50/54 con serie que empieza con NÚMERO         → Tipo+Serie+N°
# SAP:  ignorar tipos 00, 05, 14, 46 siempre
# SAP:  ignorar cualquier doc con serie que empieza con 00
TIPOS_SIRE_CON_RUC    = {'01','07','08','30','42'}
TIPOS_SIRE_SIN_RUC    = {'50','53','54'}
TIPOS_VALIDOS         = TIPOS_SIRE_CON_RUC | TIPOS_SIRE_SIN_RUC
TIPOS_SAP_IGNORAR     = {'00','0','05','5','14','46'}

def fmt_s(v):
    try: return f"{float(v):,.2f}" if v else '0.00'
    except: return '0.00'

def flt(v):
    try: return float(str(v).replace(',','') or 0)
    except: return 0.0

def norm_nro(n):
    """Elimina ceros a la izquierda del número de comprobante."""
    return str(n or '').strip().lstrip('0') or '0'

def serie_valida_sire(tipo, serie):
    """SIRE: 01/07/08/30/42 deben tener serie con letra. 50/54 con número."""
    if not serie: return False
    tipo = tipo.zfill(2)
    if tipo in TIPOS_SIRE_CON_RUC:
        return serie[0].isalpha()     # E001, F016, etc.
    if tipo in TIPOS_SIRE_SIN_RUC:
        return serie[0].isdigit()     # 0001, etc.
    return False

def serie_valida_sap(tipo, serie):
    """SAP: ignorar tipos 00/05/14/46, y series que empiezan con 00."""
    tipo  = str(tipo or '').strip().zfill(2)
    serie = str(serie or '').strip()
    if tipo in TIPOS_SAP_IGNORAR:
        return False
    if serie.startswith('00'):
        return False
    return True

def tipo_label(t):
    m = {
        '01': 'Factura',
        '07': 'Nota de Crédito',
        '08': 'Nota de Débito',
        '30': 'Tarjeta de Crédito/Débito',
        '42': 'Tarjeta Propia',
        '50': 'Dec. Única de Aduanas',
        '53': 'Mensajería / Courier',
        '54': 'Liquidación de Cobranza',
    }
    return m.get(t.zfill(2), t)

TIPO_COLOR = {'01':'#1E40AF','07':'#991B1B','08':'#92400E','30':'#166534','42':'#5B21B6','50':'#0E7490','53':'#6D28D9','54':'#065F46'}
TIPO_BG    = {'01':'#EFF6FF','07':'#FEF2F2','08':'#FFFBEB','30':'#F0FDF4','42':'#F5F3FF','50':'#ECFEFF','53':'#F5F3FF','54':'#F0FDF4'}

# ── Detectar período directamente del archivo SIRE ───────────────────────────
# Lee el período del primer registro válido del TXT (columna 2)
meses_es = ['','ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO',
             'JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']

periodo_id = None
with open(SIRE_FILE, 'r', encoding='utf-8', errors='ignore') as _f:
    for _i, _line in enumerate(_f):
        if _i == 0: continue
        _cols = _line.rstrip('\n').split('|')
        if len(_cols) < 3: continue
        _per = _cols[2].strip()
        if len(_per) == 6 and _per.isdigit():
            periodo_id = _per
            break

if not periodo_id:
    print("❌ No se pudo detectar el período del archivo SIRE"); sys.exit(1)

_anio = int(periodo_id[:4])
_mes  = int(periodo_id[4:])
periodo = f"{meses_es[_mes]} {_anio}"
hoy = datetime.now()
print(f"   Período detectado del SIRE: {periodo_id} → {periodo}")

# ── Leer SIRE TXT (parser robusto) ──────────────────────────────────────────
# Formato: fila normal=80 cols. Con salto: 14 cols (inicio) + 67 cols (resto)
with open(SIRE_FILE, 'r', encoding='utf-8', errors='ignore') as f:
    raw_lines = f.readlines()

all_rows_raw = []
i = 1
while i < len(raw_lines):
    line = raw_lines[i].rstrip('\n')
    cols = line.split('|')
    if len(cols) >= 75:
        all_rows_raw.append(cols); i += 1
    elif len(cols) < 20 and i + 1 < len(raw_lines):
        next_line = raw_lines[i + 1].rstrip('\n')
        joined_cols = (line + next_line).split('|')
        if len(joined_cols) >= 75:
            all_rows_raw.append(joined_cols); i += 2
        else:
            all_rows_raw.append(cols); i += 1
    else:
        all_rows_raw.append(cols); i += 1

sire_rows = []
for cols in all_rows_raw:
    if len(cols) < 25: continue
    tipo = cols[T_TIPO].strip()
    if tipo not in TIPOS_VALIDOS: continue
    serie = cols[T_SERIE].strip()
    if not serie_valida_sire(tipo, serie): continue
    per = cols[T_PERIODO].strip()
    if per != periodo_id: continue
    sire_rows.append(cols)

print(f"   SIRE período {periodo_id}: {len(sire_rows)} registros válidos")

# ── Leer SAP Excel ────────────────────────────────────────────────────────────
wb      = openpyxl.load_workbook(SAP_FILE, read_only=True)
ws      = wb.active if 'SAP' not in wb.sheetnames else wb['SAP']
sap_rows = [list(r) for i,r in enumerate(ws.iter_rows(values_only=True))
            if i > 0 and any(v is not None for v in r)]

print(f"   SAP: {len(sap_rows)} registros")

# ── Cruce TIPO + SERIE + NRO + RUC ───────────────────────────────────────────
def key_sap(r):
    tipo  = str(r[A_TIPO] or '').strip().zfill(2)
    serie = str(r[A_SERIE] or '').strip()
    nro   = norm_nro(r[A_NRO])
    ruc   = str(r[A_RUC] or '').strip()
    if not serie_valida_sap(tipo, serie):
        return None   # ignorar este registro del SAP
    if tipo in TIPOS_SIRE_SIN_RUC:
        return f"{tipo}|{serie}|{nro}"
    return f"{tipo}|{serie}|{nro}|{ruc}"

def key_sire(r):
    tipo  = r[T_TIPO].strip().zfill(2)
    serie = r[T_SERIE].strip()
    nro   = norm_nro(r[T_NRO])
    ruc   = r[T_RUC].strip()
    if tipo in TIPOS_SIRE_SIN_RUC:
        return f"{tipo}|{serie}|{nro}"
    return f"{tipo}|{serie}|{nro}|{ruc}"

sap_keys    = set(k for r in sap_rows if (k := key_sap(r)) is not None)
pendientes  = [r for r in sire_rows if key_sire(r) not in sap_keys]
registrados = [r for r in sire_rows if key_sire(r) in sap_keys]

total_sire = len(sire_rows)
ya_reg     = len(registrados)
pct_reg    = round(ya_reg / total_sire * 100) if total_sire else 0
pct_pend   = 100 - pct_reg

total_bi  = sum(flt(r[T_BI])    for r in pendientes)
total_igv = sum(flt(r[T_IGV])   for r in pendientes)
total_cp  = sum(flt(r[T_TOTAL]) for r in pendientes)

print(f"   ✅ Registrados: {ya_reg} ({pct_reg}%)")
print(f"   ❌ Pendientes:  {len(pendientes)} ({pct_pend}%)")

# Fechas del período
fechas_txt = [r[T_FECHA].strip() for r in sire_rows if r[T_FECHA].strip()]
fecha_min  = min(fechas_txt) if fechas_txt else '-'
fecha_max  = max(fechas_txt) if fechas_txt else '-'

# ── Agrupaciones ─────────────────────────────────────────────────────────────
by_tipo = defaultdict(list)
for r in pendientes:
    by_tipo[r[T_TIPO].strip().zfill(2)].append(r)

by_prov = defaultdict(lambda: {'count':0,'total':0,'igv':0,'tipos':set()})
for r in pendientes:
    razon = r[T_RAZON].strip()
    for s in [' SOCIEDAD ANONIMA CERRADA',' SOCIEDAD COMERCIAL DE RESPONSABILIDAD LIMITADA',
              ' EMPRESA INDIVIDUAL DE RESPONSABILIDAD LIMITADA',
              ' S.A.C.',' SAC',' S.A.',' SA',' E.I.R.L.',' S.R.L.',' SRL']:
        razon = razon.replace(s,'')
    if len(razon) > 42: razon = razon[:42]+'…'
    by_prov[razon]['count']  += 1
    by_prov[razon]['total']  += flt(r[T_TOTAL])
    by_prov[razon]['igv']    += flt(r[T_IGV])
    by_prov[razon]['tipos'].add(r[T_TIPO].strip().zfill(2))

top10_count  = sorted(by_prov.items(), key=lambda x: x[1]['count'], reverse=True)[:10]
top10_amount = sorted(by_prov.items(), key=lambda x: x[1]['total'], reverse=True)[:10]
num_prov = len(by_prov)

# ── JS data ───────────────────────────────────────────────────────────────────
# ── Ordenar pendientes alfabéticamente por razón social ───────────────────────
pendientes.sort(key=lambda r: str(r[T_RAZON] or '').strip().upper())

js_data = [{'fecha':r[T_FECHA].strip(),'tipo':r[T_TIPO].strip().zfill(2),
    'serie':r[T_SERIE].strip(),'nro':r[T_NRO].strip(),'ruc':r[T_RUC].strip(),
    'proveedor':r[T_RAZON].strip(),'bi':flt(r[T_BI]),'igv':flt(r[T_IGV]),
    'total':flt(r[T_TOTAL]),'moneda':r[T_MONEDA].strip(),
    'orig_fecha':r[T_ORIG_FECHA].strip() if len(r)>T_ORIG_FECHA else '',
    'orig_tipo': r[T_ORIG_TIPO].strip().zfill(2) if len(r)>T_ORIG_TIPO and r[T_ORIG_TIPO].strip() else '',
    'orig_serie':r[T_ORIG_SERIE].strip() if len(r)>T_ORIG_SERIE else '',
    'orig_nro':  r[T_ORIG_NRO].strip() if len(r)>T_ORIG_NRO else '',
    } for r in pendientes]

unique_dates   = sorted(set(d['fecha']  for d in js_data))
unique_monedas = sorted(set(d['moneda'] for d in js_data))
unique_tipos   = sorted(set(d['tipo']   for d in js_data))
razones        = sorted(set(d['proveedor'] for d in js_data if d['proveedor']))

date_opts  = '\n'.join(f'<option value="{d}">{d}</option>'             for d in unique_dates)
mon_opts   = '\n'.join(f'<option value="{m}">{m}</option>'             for m in unique_monedas)
tipo_opts  = '\n'.join(f'<option value="{t}">{tipo_label(t)}</option>' for t in unique_tipos)
razon_opts = '\n'.join(f'<option value="{r}"></option>'                for r in razones)

# Tarjetas por tipo
tipo_cards_html = ''
for t in ['01','07','08','30','42','50','53','54']:
    rows_t = by_tipo.get(t, [])
    if not rows_t: continue
    tot = sum(flt(r[T_TOTAL]) for r in rows_t)
    col = TIPO_COLOR.get(t,'#374151')
    bg  = TIPO_BG.get(t,'#F2F4F7')
    tipo_cards_html += f"""<div class="tipo-card" style="border-left-color:{col}">
      <div class="tipo-card-lbl">{tipo_label(t)}</div>
      <div class="tipo-card-num" style="color:{col}">{len(rows_t)}</div>
      <div class="tipo-card-total">{tot:,.2f}</div>
    </div>"""

# Top 10 HTML
def tipo_badge(t):
    col = TIPO_COLOR.get(t,'#64748b'); bg = TIPO_BG.get(t,'#f5f7fb')
    return f'<span style="background:{bg};color:{col};border:1px solid {col}33;border-radius:20px;padding:1px 8px;font-size:9.5px;font-weight:700">{tipo_label(t)}</span>'

def make_prov_rows(top_list):
    return ''.join(f"""<tr>
  <td class="proveedor-cell">{n}</td>
  <td style="text-align:left;padding-left:13px">{''.join(tipo_badge(t) for t in sorted(d['tipos']))}</td>
  <td class="num">{d['count']}</td>
  <td class="num money">S/ {d['total']:,.2f}</td>
  <td class="num igv">S/ {d['igv']:,.2f}</td>
</tr>""" for n,d in top_list)

prov_html_count  = make_prov_rows(top10_count)
prov_html_amount = make_prov_rows(top10_amount)

now_str = datetime.now().strftime('%d/%m/%Y %H:%M')
LOGO_B64 = 'iVBORw0KGgoAAAANSUhEUgAAA6oAAAHnCAIAAAA+RzfSAAAQAElEQVR4Aey9TWwex53nb1sjmhrJEiV5PPIlojfrhQZeQOQkQOxDLApxANtzMHUYLwzMQuQsMAudRJ52chJ1SuYk6mRkgP9Swg5grHMgdRjbQLwQ5RziANmQAlYYYb2JyVysOHqhbCpiqNj6f+Sy283ql6e6+r2fr1HpVFdX/er3+1R117frefjokfv6TwREQAREQAREQAREQAT6hsAjD+k/ERABEehTAgpbBERABESgHwlI/vbjqCtmERABERABERCB/ibQ19FL/vb18Ct4ERABERABERABEeg3ApK//TbiilcEthLQmQiIgAiIgAj0GQHJ3z4bcIUrAiIgAiIgAiJgCOjYrwQkf/t15BW3CIiACIiACIiACPQlAcnfvhx2Bb2VgM5EQAREQAREQAT6h4Dkb/+MtSIVAREQAREQAZuAzkWgDwlI/vbhoCtkERABERABERABEehfApK//Tv2WyPXmQiIgAiIgAiIgAj0BQHJ374YZgUpAiIgAiKQTEBXREAE+ouA5G9/jbeiFQEREAEREAEREIE+JyD5G5oAyoqACIiACIiACIiACHSdgORv10dY8YmACIiACwHVEQEREIG+ISD52zdDrUBFQAREQAREQAREQAQeesiWv2IiAiIgAiIgAiIgAiIgAh0mIPnb4cFVaCIgAtkIqLYIiIAIiEA/EJD87YdRVowiIAIiIAIiIAIikEagr65J/vbVcCtYERABERABERABEeh3ApK//T4DFL8IbCWgMxEQAREQARHoOAHJ344PsMITAREQAREQARFwI6Ba/UJA8rdfRlpxioAIiIAIiIAIiIAIQEDyFwhKIrCVgM5EQAREQAREQAS6S0Dyt7tjq8hEQAREQAREICsB1ReBPiAg+dsHg6wQRUAEREAEREAEREAEviIg+fsVCf3/VgI6EwEREAEREAEREIFOEpD87eSwKigREAEREAF/AmopAiLQbQKSv90eX0UnAiIgAiIgAiIgAiKwhYDk7xYcW090JgIiIAIiIAIiIAIi0DUCkr9dG1HFIwIiIAJFEJANERABEegsAcnfzg6tAhMBERABERABERABEYgS6CV/oy1UIgIiIAIiIAIiIAIiIAKtJSD529qhk+MiIAJlE5B9ERABERCBLhKQ/O3iqComERABERABERABEchDoNNtJX87PbwKTgREQAREQAREQAREYCsByd+tPHQmAiKwlYDOREAEREAERKBjBCR/OzagCkcEREAEREAERKAYArLSVQKSv10dWcUlAiIgAiIgAiIgAiIQQ0DyNwaKikRgKwGdiYAIiIAIiIAIdIeA5G93xlKRiIAIiIAIiEDRBGRPBDpIQPK3g4OqkERABERABERABERABJIISP4mkVH5VgI6EwEREAEREAEREIFOEJD87cQwKggREAEREIHyCMiyCIhAtwhI/nZrPBWNCIiACIiACIiACIhAKgHJ31Q8Wy/qTAREQAREQAREQAREoO0EJH/bPoLyXwREQASqIKA+REAERKAzBCR/OzOUCkQEREAEREAEREAERKA3gazyt7dF1RABERABERABERABERCBxhKQ/G3s0MgxERCBphGQPyIgAiIgAl0gIPnbhVFUDCIgAiIgAiIgAiJQJoFO2Zb87dRwKhgREAEREAEREAEREIF0ApK/6Xx0VQREYCsBnYmACIiACIhAywlI/rZ8AOW+CIiACIiACIhANQTUS1cISP52ZSQVhwiIgAiIgAiIgAiIgAMByV8HSKoiAlsJ6EwEREAEREAERKC9BCR/2zt28lwEREAEREAEqiag/kSgAwQkfzswiApBBERABERABERABETAlYDkrysp1dtKQGciIAIiIAIiIAIi0EoCkr+tHDY5LQIiIAIiUB8B9SwCItBuApK/7R4/eS8CIiACIiACIiACIpCJgORvJlxbK+tMBERABERABERABESgbQQkf9s2YvJXBERABJpAQD6IgAiIQGsJSP62dujkuAiIgAiIgAiIgAiIQHYCeeVv9h7VQgREQAREQAREQAREQARqIyD5Wxt6dSwCItB2AvJfBERABESgjQQkf9s4avJZBERABERABERABOok0Oq+JX9bPXxyXgREQAREQAREQAREIBsByd9svFRbBERgKwGdiYAIiIAIiEDLCEj+tmzA5K4IiIAIiIAIiEAzCMiLthKQ/G3ryMlvERABERABERABERABDwKSvx7Q1EQEthLQmQiIgAiIgAiIQHsISP62Z6zkqQiIgAiIgAg0jYD8EYEWEpD8beGgObv8+drG5uKKSXfPLX06c9E9bSxcNQ05OneoiiIgAiIgAiIgAiLQdAKSv00foZ7+GY1r1O3axPyNsTnSRw+fIv1u7w9vHJ0zaW1yYf30onu6dewN05AjpkjXhn6IZdInU28jo40+DtxTRgREQAREQAREQARaQUDytxXD9LWTn62ssR2L+jRKF1VqNK5Rt3fPL29eWiF93aC43P3bG1gm3Tn7PjLa6GMcMLIYTXxn9uf4hhwvrk9ZEgEREIEWEJCLIiAC7SIg+dv08UJNoinRu2y7IjQ/fuoM27GoT6N0m+C9kcVo4k+m38E35PjHw2dujb+Bz3jeBA/lgwiIgAiIgAiIgAgEBCR/AxT5M4VZYIv37rkl9nevj7yOmkRTonfZdkVoFtZHmYY+W13buHB1/fQinrM9jHBnb3hj4SpSvsxuZVsEREAEREAEREAEehOQ/O3NqJoaSEMEIjKRrVO2eNcmF9jfvXf5WjW9l9oLwp294VvH3kDKI+i1K1wqbRkXgZoIqFsREAERaA0Byd+ah4qN3juzP2d/FGmIQEQmsnVas09ldo+gN7vC14Z+eGv8DTa50f1ldijbIiACIiACIiACIrCFQNHyd4txnSQSMKqXrVA2ej+Zfof90cSqHb1w//bGxoWrbHI/0P3SwR0dZYUlAiIgAiIgAg0kIPlb6aBYqpet0Eq7b2pn0sFNHRn5lZGAqouACIiACLSBgORvFaPE5/t8yn9jbM7s9Ur1JkEPdPDaxPzGwtWkaioXAREQAREQARFoFoFWeSP5W+5w3Vu+hpL7ePgMn/L34TccvOHePb9869gbcPt05iJb5t521FAEREAEREAEREAELAKSvxaQwk7Z7r0+8vr10ddRcvdvbxRmt58Mfba6tn56kS3zW+Nv6CeE2zLy8lMEREAEREAEGk5A8rfgAWKrkg3La0M/ZLtXX3IoCu7Ghas3js6xGcxLxedrepcoiqvsiIAIiIAIFElAttpCQPK3sJFC+K5NzLNVuX56Udu9hWENGWIzmJcKRDAvGBLBITDKioAIiIAIiIAIZCAg+ZsBVlJVPpc3wvfu+eWkOiovigCvFrxgIIJhzitHUWYLtSNjIiACIiACIiACzSUg+ZtrbBC+N8bm+FxewjcXx+yNEcEwZ69dIjg7PLUQAREQgfIIyLIItICA5K/nILHvaISvfs/Bk2BBzYwI/mTqbX0doiCiMiMCIiACIiACHScg+Zt5gBG+7Diy7yjhm8au2mt3zr7/8fAZfSe4WurqTQREQAREQARaSUDyN8Owsb+IwEL4suOYoZmqVkLg/u0N853gu+eWKulQnYiACIhAAgEVi4AINJuA5K/r+CCq2F9EYLk2UL06CCCC1yYXro+8vrm4Ukf/6lMEREAEREAERKDpBCR/e48QQgo5hahCWvWuHaqhbF0E7l2+duPo3K3xNz5bWavLB/UrAiIgAiIgAiLQTAKSv2nj8vnaxtrEPEIKOZVWT9caSWDjwtXfj7z+6czFRnonp0Sg4wQUngiIgAg0loDkb+LQmG876Gu+iYDacIEN+/XTi9f1XYg2DJZ8FAEREAEREIFqCJQtf6uJouBe+MT8xticvu1QMNb6zLF5zxa+fhytvhFQzyIgAiIgAiLQIAKSv/Zg3Jn9OZ+Y60fNbC7tP79z9n1tA7d/GNsVgbwVAREQARFoIgHJ369HxWz6fjL9Dp+Yf12qXIcIfLa6pm3gDo2nQhEBERABEWgsgUY7Jvn75fBsLDz4Mylt+n6Jo9P/xzbwzbG5e8vXOh2lghMBERABERABEYgnIPn7kPl5h1vH3tCmb/wc6WLpvcvXro++fmf2510MrmExyR0REAEREAERaBiBfpe/bAGyEaifd2jYtKzInU+m37k1/gbvPxX1p25EQAREQAT6ioCCbSqBvpa/d88t3eBD8Mv6ELyp07N8vzYuXL0+8jpvQeV3pR5EQAREQAREQAQaQaBP5S8bfmsT8/pps0bMwbqd+Gx1rewvQtQdovoXAREQAREQARH4mkA/yt/PVtY6/4WHbQeHBo4MD75yaNepMdLQ3Pj+i5MmPfHh9JP3Tyelx5dOmGocd595kbYkTJEe3jP49cTpXO6T6Xd4I+K9qHORKSAREAERqJOA+haBBhLoO/nLx9y/58Pubn3hAW268+Sz6FQ0q1G3T6xM71+c3Lvw2mMzR0k7JkYHxoZN2jY8lDIRt48cMNU47px6jrYkTJEOrP0A0Yw+3jv/Gn3tOD6y/fCBFFOtu3T3/DLvRbwdtc5zOSwCIiACIiACIuBOoL/k791zS3zM3YFfeDB6FxlqxC7adPfsS+hUNGu6unWfGUk10ceD44foa+jcsceXTxhBzO5yLzWcZK9Z5fcuX3vwdqTfRGvWsMgbERABERABESiSQB/J30+m3l6bXCgSXrW2kLzsubK/i+I0ehcZWrbYdQkRQczuslHDf3nrB4hytqLbuzHM2xHvSLwpucSuOiIgAiLgQEBVREAEmkWgL+Tv52sbaxPzd86+3yz2Dt5sOziElERQGsnLniv7uw7taqvyyNAgopytaDaGkcLsCg++cqiNXxrmTemOfhW4tnmkjkVABERABESgRALdl79o35tjc3fPL5dI0dG0czWjeh9fOvHEyjRSEkHp3LRBFZHC7ArvXXjtwNqDLeEdx0fapYPNH8M1CKhcEQEREAEREAERKIJAx+XvZ1/8yMO9lvyhm6V6t4905w/LUPBD5461Tgfz1sTnBrxBFXGvyYYI9DcBRS8CIiACjSHQZfnboh95YGd07/xrZq+3S6o3Os+NDibSobnxgSPD0QpNK0EB3xybkwJu2rjIHxEQAREQARHwJlC1/PV2NGtDtO+Nsbn7tzeyNqyyPtu9u8+8+OA7sueOoQur7Lrevsz3IvYvTj7x4fTOk882/EsRfHogBVzvhFHvIiACIiACIlAggW7K3+ZrXzY+zXbvzqnn0IIFjmi7TG0bHto9+5LZDOZloLHOSwE3dmja5Zi8FQEREAERaAKBDsrfhmvfHcdHHl86wcZnX233ps91XgB2TIwignkl4MUgvXJdV6WA6yKvfkVABERABDpAoFEhdE3+Nln7Inz5rH/o3LFuf7s3z/zmlYAXg/0XJ5spglHAt8bfyBOg2oqACIiACIiACNROoFPyt7HaNxC+fNZf+5A334GBseFmiuCHjVUyswAAEABJREFU9wzunn2p+QCb7aG8EwEREAEREIGaCXRH/jZT+w6+csjs+Er4Zp3pgQhuyD8gh/ZFlGvnPus4qr4IiIAIiMBXBPT/TSHQEfnbQO2LaOND/L0Lr0n45pnsiODHl08MzY2jPvPYydmW3qV9czJUcxEQAREQARFoCIEuyN+maV+kEnIN0YZ0a8gwp7ixsrKScrUhl8wfxu06NVaLPwxowdq3ljDUqQiIgAiIgAiIwBcEWi9/P1/buDX+RnN+33fnyWefWJlGrn2Bt+mHiYmJkZGRxcXFpjv60EOPDA0+NnP0iQ+nK/6rOGnf5s8NeSgCItAmAvJVBBpAoN3yF+17c2zus9W1BpB8yHzbYffsSwi1JvjT04eJiYnz58/fvn376NGj586d61m/CRW2DQ+xEcvmOqq0An/ohe70fd8KUKsLERABERABEaiMQLvlL/u+9y5fqwxWSkd8Lt+WbzsQxdraGpu+aF/yJk1OTlasgE2/fscdX/xI8OArh/yaO7aS9nUEpWoiIAIiIAIi0C4CLZa/axPzm5fq/94qm76PL53gc/m2DDzad2xs7PLly5bD7VLAbLHvXXht7/xriFQrkEJOMat930JIyogIiECUgEpEQATqJdBW+Xtn9ud3zy/Xy47ed558lk3fFn04nqR9iYWEAp6YmCDTljQ4fuiJleK/DSzt25YJID9FQAREQAREwINAK+XvxsLVT6bf8Yi2wCYPFNLFyd25/hGEAt1xMrW8vDwWt+8bbnz+/Pl2KWC2gdmm3X3mxXAUefIPRnZxskWvNHmCVVsREAEREAER6EMC7ZO/95avrU3M1ztUA0eGH2w6jg3X60am3l20rzHYOgWM2zunnnt86cS2g0Pk8yRp3zz01FYEshBQXREQARGojUDL5O/nDfiZs12nxthuZNOxtkHL3rHRvrdv33Zsev78+fHx8bW1Ncf6TajGfu3jyyfy/D2ctG8TxlE+iIAIiIAIiEDZBOqWvxnjuzX+Ro0/c4Y82jv/Wov+ys3QXVxcHBsbu33bVfuaVhcuXKBVuxQw7yR7F17j/cSEkOnI4PJWg4bO1EqVRUAEREAEREAEWkegTfL305mLNf7Uw/bDB5BHg+OH2jXG586dO3r0aFbta2K8fPly6xQwnvN+wlsKcpa8Y6Iygyvt64hL1QojIEMiIAIiIAJ1EGiN/N1cXFk/Xds/TjZwZHhfC/8cCu07OTmZZ14ZBbyyspLHSPVteUtBzjp+Fbjz2nd5eZmZMDMzw8sM6eGE/7g0Pj5OtYWFhRaNuIluamoK/0kJwT3MJdLMzAwoaFL9nHTsEd9wkoEYGRlJiqX6ckfnC6/Gp0+Li4sAmZiYYPiGh4djYx8aGuIq0KjJ7KVV4Z4UYhDHcM/M1eGEWGIDLLUQdIVEJyMikJlArQ3aIX8/X9u4Of5GXaB2HB9BS/HBel0O+PXLMp9T+5p+UcCsxKzK5rQtR7ZyH18+wZ59usNd1b4stEwABAHKYHR0lJlw+vTpS1/8lwSEixcuXKDasWPHnnrqKZZnNAerdVL9Gsuj0Z09exb/SUlecYlEdKAACFiAAyJMJTWpspz3DVQRXuEbTjIQ3HdVOtCovnjaIGR57Ozdu/fo0aMAOX/+PMO3uroa6yefbnEVaKdPn2b20orZC8/mzF48Yb7hGO6ZuZoUS2yAKhQBESicQDvk763xN+7f3ig8eBeDu8+8OHTumEvNRtVBuLDMF+USqws7BKxJRRmsxg5vLOzZ8/aS1F0ntS+7ZWahZQIgCBi7pPDTy1me0Rys1kNDQygJ9Fl6fe+rmRoGMiJndGABDkZQJNwsQMvkRoGVAYsDTz31FKoIrwq03DpTvIrMzs6iXM07QJ4XAGYvPJm9WKt39jJj8QFPmG+tGxE5LAIdJtAC+VvjV36H5sZ3Tj3XuuFnNUW4FOs2C3NLFTBvL7EKuHval71MFlp2y4pdaBl6lAT6jAlQo0w00ZUhI7hZgAY6uij2rulpjT1OwOJAz5rdroDwRaQyBNPT0yjXAoPFmpm9PBV50yjQck9TdMctw4zFh56VVaFvCSjwugg0Xf7eW75W11d+0b47JkbrGhjvfnnKl7SaIoPYlaleInijCBpGFXDHtC+qFOnAXmapCy2fLyMTWdEr/hzA7J+VHR3o6GJkZASYwcwpL4M2oi8+rC+vi1ZYRvjyDsDsRaTyhCnPZ56KvGlMTExAvrxeAstMWsaXWyYoUUYERKBRBBotfz9f27hdx79wgTZ6fOlE67QvCwmfevOUL3WGIRFaqoB5nzFkGN/9zfpDRuOXz9EMOqoU9ebTPnsbVnTegtiro+vsrbO1QKmgtqvcP+MDd2ByH5UaHe8PaCP6yoajc7WNRuQdoFThG8bG4xHys7Oz4cLC8wh6Jm1lQRXuvwyKQD8QaLT8XZ+5eO/ytYqHoaXaiNUaoVDsp95J5FuqgHmfQQG3dHxjx2JhYYFts2oG3XKAvTpkxOLiolVe4ClvWXSB2i7QpqMpkAIWvI71M1VD+3Kr9rk24nnFRiwasbLXtmCMID89Pc0QrKysBIUFZogLQV+gwT4wpRBFoAYCzZW/m4srd86+XzGSlmoj1hKe5lVuJk1OTs7MzFQ8Ovm7QwE/sTK9feRAflO1W2CVRT2wltflyerqKhulbAMX7gDzmeiYYzVGR9fgLTw6aV9mi4HARiz5uhKvVbxcFf6Gw7ytN666eKpfEWgdgYbK38/XNtYq/9pDS7UvexgVa18zy0+fPs2z3uQLOFZl4pGhwaq6Kqsf1CErd0NWWbMNjEtFRYsp5nNzosMZXCokOiP7ENaFWGupERQnSKt8V08CxUDwhlPga/zs7GxD5m1SyCoXAREICDRU/t6Z/Xn1/7jx0LljrdsXZEFFCdW1lvCsb6MCDmZ/GzOMeEPUQ0CP6cckxLGgxDuDkeHhYQx6Wyi8IduEAM+vgLEwPj6O5CrcwxYZPHfuHIqzURCKeo1fXFycnp72Hwu1FAERqJZAE+VvLb/2MDQ3Pjh+qFr4eXtDK7Aw17uWGAW8traWNxi1dyBgRrxR6tB4vbq6ylTEPXPqd6Q5Ruqdz7GeA3xkZAT3Yq86FvKiCCXHyp2sdu7cucnJXP8IZUlYzEMsj3EegLzb5LGgtiIgAhUTaKL8/WTq7YopoH13NOA3zjJFzWLcEK3A4oEnLACZ/FflrASaM+KxniNbmQY4GXu1ZyENaY6RnjVrqYByxT3vSc7W4IULF2rxvCGdNlb7Gj48xHg/MXmP49TUVGOnrkc4aiIC/UCgcfL37rmlzUul/EFu0nDuPPls67Qva8no6GhzHrhsj+URB0lDo/KAAMKL7aXmjHjgWDiDe0yDlZWVcKFLniY0pLlL5brq4B5OMhAeDuSRVh7dNa0JzyuXfd963fZWwLy50bZe59W7CIhAVgLNkr+fr23cnnonawx56u84PrJ79qU8Fqpv28y1RAq4vJmA5EJ4sQFZXhdFWUYjItNx2N0glWlCQ/cmddU0kzxr79ywrRi7rHE51mfnu/na18SCip2dnTV59+PMzIx7ZdUUARFoCIFmyd/1mYv3b29Uhmb74QPSvgXSRhzk/4pkgf50xhQfrcK2LeHgaqbNTrQvTboaHXF5KCpadSOtrKwwvi2KZXp6Gr3u7jAB9vnXWtxZqaYINIpAg+TvveVrdyr8od+H9wzuW5xs189gsc3Q8H2U1dVV9imXl5cbNctb7QziiU2pdoWAIMBtF5+Z0pcuXXKp2Zw6DAcbuo7+cC9kEfeOVltTDe3bin39MFB85hOJcElK3n0mpBjRJREQgeoJNEj+VvwXb/sXW6Z92VE7ffp09VMka4+sdlLAWaEl1Uc8IRCTrja5HLfZGEv3kOhaMaWjUbAf3zM602phYcFk+vDIHGij9OcJhgJ2HK9+Hl9HRKomAs4EKq3YFPm7ubhS5V+8Dc2Nt+snftG+7DlVOjVydMb6gQLWwpAD4ZdNkVnA/PKkoP87ePDgK6+8ciry35EjR/bs2VNQJw/hNpM23VrPCunNU64SCOGQyKRU877kEp0xnumTdNOkG8dS320OHz7M4JJKYsUnEi4fX/AK1EZ9XxI0mRWBdhFoivyt8t9423F8ZEd7fuaMj+FQCS3SvuYGQB8cO3ZMnwwaGn5H6LEM+7WNtkIxnDlz5sMPP2TN5s2EnTkrIdSYbEtLSwhjJHLUQtYSnKejpFb0Hi8dkhr0KkfTz83N4f/9+/cJhHBIZDilkNip0MtGhuvp0QWGqBbks2YYsvn5+Vu3bhFCvSmr59TnzY1jUYkJefLkyYAG2prBJRksFy9eZNIePny4qO6ww/xk8pBJSbiRcrXnpePHjzMz79f9Hxh7uqoKItA9Ao2Qv3fPLVX2b7y168/deP6yjdo67RvcJ5OTk2i44FQZdwIMfVEC4vDhw+gDlmoMDg8Pp/swMjLCwr+ysoKURHOkV+55lR5j6xCdy+5abFurkP1dpA8aEanNiyL+WxU4pRBPqEA1KtOEwvwJm+lGYJ5eIeUqYp3mfAo/NDSUUq2Zl7jr8+j+cFBs8aJ6mZBMmCQaPCSZtOBCTR4/fjzc3DvPCzw205vTY3qFlKvcX1BiZqbU0aV+JKCYqyLQCPn76cxiNfE+vGdwz7ljbflzNyQCj/Vid8iq4RzuBQXMuhUuUd6FANBYgF1qptdh15NFmomUXi16FSlJw5MnT0YvOZagMjESW3lmZqaQ6HAPYYQ1R41INSrTBHEZ61imwtXVVRRMShNu4ZSrKZdAl245pW0TLgE5vxtAQCOyN4nqdbSGmoQbL3u88jk2Sal29uxZpkpKhfSrKQ3R9Em3RkorXRIBESiQQP3y926FW7+PzYy15Su/LJxIlrZrXzNTp6en9aw3KByPjD7y17FyUjXUA5thPXcok5pTjljEDSQIpjh1T9Rnk3VlZSVWBhEdGsXdWmxNukDl4B5OxlZIKaQJO8EecUVtxgYYVINAkM+U4d7HyUxNmlOZweXFIKc/6Ffo+T03oLe4uFjINnBJ45vnrswJVs1FQAQMgfrlb2Vbv4OvHNo59ZwJu+FHdt3YxuiG9jWoz58/77eSmeb9dkRA5NwcNeqBWZQfHQOHmEBuupiiWiB8kwQcmjVndAcPHsQlVI6LS0l1MsWVZGQ1dQMYAZfUML28kIFL76K8q+mS0aVflCvPwKT542KBttxEfPThUjmlDq9JvK2lVPC71PM7SH5mO9NKgYhABQRqlr93q9r6NV97qABo/i547rOus6zmN9UoC1LA7sOBQHSvHK2J9kUdogCil/xKUGOIiZ5tjWpB/aR37WIqpS+0L/cILqXUcbyEEUAh2R3rx1bLGU6szfYWwjPns4tZVBRSNlnZ488Dk/e0opwJu8HEC58qLwIiUD2BmuVvZVu/Qy35yi/rOtqXZ271U6GCHlHAROe2m1KBOw3tgjmQR0Ag5tiyShegHpGPj4+n7KUhWSq7HegAABAASURBVD788EOEQs9tLXxrVHQIERSbB5CgyaVLl1ZWVoLTPs8wB/IQOHLkSE4LVu/s8Z88edIqzHRarD+ma00Yw0FHEaiRQJ3y925VW7+DrxwaHD9UI2XHrlEGqMOual8DAa1AjFLAhkbsMedyyyzqqUFj++1ZyF4a6sSqRomj8DUNcc9k/I40HxkZ8Wub1AqDOfcI8SrWOFM9trxn4crKSs86zayQhMLFW/b18zRP6oLPUpilSVd7ll++fLnw4XAw2NMvVRABEchFoE75W83Wb1u+9oDoOXbsWLe1r5mqLCfIAilgQyN6zKMA2OiCbdRmUSXM0sAUkuLixYtsnQ4PDweFPTPNjI49wjy/BRHG0pOASwWoulRrWh3czvMEA2Phn1oYRMw6PhUxeY8jzWNbed9rSQZje1GhCIhAGQRqk78bC1er+a3fx2bGmv9LZzz3Jycn7QHu7rlRwHzK390QPSNjW8j7uwEs8DMzM54duzVD6Z46dSoQvlmXf0bcWx6VHR17hHThhsGuxXyOfZ1jX9mu6nbOHMAft7oNqpVH1R0/fjzrdHKPHFWd59ZIiguz7j6Ea/LA514IlygvAiJQMYHa5O+d2Z9XEOrAkeGdjf+1B9a5vtK+ZtxRDKx2WgMMjeDI/lmQz5qZmpryXo/d+0JG4CRj594kqEnDIJ81w21SanQoewBm9SqoHxsaDvOBflAnU2Z6ehrUsao6k50qK+e5nQnWdrXQcwbXeywuXboU64vfXYApXgJpiwgmryQCIlALgXrk773la5uXqvhy2+7Zl2rB6t4pn7qyzrnXL7Ym211s5hVr092aWQPyLJnufbWlZh4aLPANDzNWI7r4zEQdHx93qZmnzsTEhHfzpIEbGxvztnn69Om9e/diobwEVXQnOoyPHbz9DBomycSgQlKGrV9eP5KuFlVOpN6mYseX3X1mpp/N27dvs+vBC1J5g4tlpjRRc9+16z3KD6laiUAmAvXIX/et30zBWJV3HB9p+D9ywbPp/PnzltuVnR4+fJg1j4djzr/7yeMwawDP6NilJY/Z9rb1RoGAYClteODMNz8PUWkVRIcC8/4GMAojNjQ8jy13L7xU5n8XLlw4ffo0Ouypp54ifN6gvGegd0NQ5KeEkZ4pTy9J0eWxicM8AMsc3kvnz59nfI8ePcp7FGKdj1Ckg8GuJAIQqEH+fr62cff8Mn2Xmh7eM9jwrd/atS8LtpEUeFKvAh4dHWX/qdT50BbjrIV+ruZchv06zdrq8uXLWZuY+pVF591RkqrAoPcG4UMPPWTCr+a4urp69uxZbkbeSHk4ZO3U+90GPlDK2p1HfR533q83SdHx8PTwpJYm3H3T09PoYHxOCqcWx9SpCNRCoAb5e/fcUgWh7px6trF/8cZKyQLDe3kFHGK7YN+X5Y3FILjKA/HixYusQ0FJxRn2n6SA8zBnRuVpXkHblZUV716qkUe4540RbUHz2MSWamx5Ywt5Bzt69CgoMg1Z0v5ozzDpqGedoip495UUHQa9v1JcVFBZ7bD0PPXUU3zux0qUta3qi0CVBErtqwb5e2f2/VJDwvi2g0M7m/oXbzxxeGKywOBnLYn9D0v7GjfwinIpYEOjlmPSEtvTGRbg8MtMz/q1VMikpcIeHjlyJHxaan54eNj7FuDWjvUN+ettM9ZgNYU8o/i4vIKXUnqpJiJ68e4raXCxiY7k2Lp0+vRpnvnez5zWxSuHRcAiULX83VxcqeD3zhr7Y2c8Q3nipGwUWcNT+Onx48cXFhaSpBJrQ+0KGK1QeNRtMcj08HMV0ebXsA2tHkqariU5z13gZ3l5eTm2If63dFbfvn2bj2X4aCg2LqswKXyrWvSUR2K0sKQS78FN8Qc+Vb6hpXiS9RIrEfAreMPJ6pjqi0AFBKqWv38o/5sPbP3umBitgF3WLlgeeNbwxMnasKj6aN+eTzqWBxTw4cOHi+o0q52zZ8+ynGRt1ef1GbUOE+hAdGwQ1nhP5ZwbfFY+Pj7e892sZ4WcbhTSnFcRPzuXLl1KaTg7O5tytcmXzBtOz3WhySF03TfFVxaBSuXv52sbGwtXywrlK7ts/X6VbdD/1659T5486fiMGxkZqVcBs9xOTEw0aPAa74r3ol5lZNwCVXbn3dfY2Jh325SG3H1t/AqEiejChQsoYJPXMUqAZ2aNvyAZ9SdrCXv8zM+srVRfBFpNoFL5+8eFf7t/e6NUXs3c+mXhZ03lPbvU2FOMz83Nzc7OplSwLiGnmqCAW7GfZKEr5bQTRvt8NFFIme7Bpo052596KU0ZFDb4X3nllZQKDb8kBdzwAZJ7hROoVP7+4Vz8d+MKjKqBW79N0L4e65ZRwDV+p409YN4Z+lwzFXhr1G4K/Ve7Dy4OrKysuFTzqMNt2Oo9Qm5JRJ5H4H3ShA3U9n7FhTFCAbNakWlckkMiUAKB6uTvZytrZf9Lbw/vGWzat355II6Ojta778ui6zdzjAI+fvy4X/P8rS5fviwF7IKxFYsW08kllmidil+BypO/hIZ8rPGGwoGc6fTp04uLizmNdLU5Mxw4rVbA4+PjFd9uXZ0Miqv5BKqTvxsL/1Y2jp1Tz5bdRSb7aF/epzM1KbDynj17Ll686K19A0+IosYFO1DAgT/KRAl0e8WqWNyXDZMbqtV7wEkKaWRkJDozXUpKfd+wHCh7LrVdAa+uruZfMizmOhWBZhKoTv7ePVf6Nx92Num3flnk6tW+i4uLbJ0WMu2IpV4FPDw8XPa6VQionEYI08/CpUtpf5buZ7M5raqUR0TNGxfHUtPMzMzc3BwvqKX2UpJxPsvC/6hxlF+00KWkylu7grkEByIq/IHpQrKQOhcuXGDtKMSUjIhAkwlUJH8/W1m7d/laqSB2HB9pzj/zNjU1Vbv29d6MiR0mFPCZM2diL1VQyIqLlGdRqaCvGrvwlr/43Hw4jCB+eiR2pCpQLcaxyhZ+9tjoq6UflJ89ezY6Isg+wzDrEQ5Zm3jXr6wvHpjtfcOJfb3xZq6GItBMAhXJ30q++fBcQxCzsLE8lOCMk0kWVB7xxWpf0zGange6yVd/7BMFfPDgQT+2DLpfwypbee93LiwsVONnZR0RDjcpLy28VXpjwUhdKaqQCMfPGbbbo2Laz1TPVlWOLwsBcbVxG5hPk1rxPOk53KogAikEKpK/m4srKU7kv7T98IHtIwfy28lvgUfe+fPn89vxs1Ce9jX+EF3tCphtFeNMJ4/eGqIVP6rlHV1lg16lPDITmLdKRBIi2PvNx9ip+Ago60vS3oOL51jj2Cvlvc7LBp8k5LWSpf3Q0BBT98MPP0QEt+slpxXPkyxDoboiYBOoQv4++NcuLly1ey70fGcD/uiNxWB8fLx27csD96Ey/0MBz8/P1/UoZw94cnKSFaXMEOu07a0hWNebv2HjHR0bhGiXsgcGEQbGsnuJ2ueeNSJ4aWnp5MmTNf7aYNS3pBLuRHCFrxKFt4KfnZ0NmyopX00vUeeHh4d5ZK2srLB3cPz4cW9KUcvllVy4cIEVrTz7siwCtROoQv4WufUbB+zhPYOPjv9V3JXqynhSjI2N8ciorsutPbFkon6Ghoa2Fpdyhsqnr7oUMCF1WAEziwjQL83MzPg1zNQKGcroZ2oSVG54dHXJo4APrwf4AN779+8jhS+W+R/7zTn/mQb8DDw3Gfw3maxH3joQiFlbZaqP+qxxbwJXeTizd0CYeML4ljm2F+fn50+dOpVTZ0fHlyiURKAzBKqQv2V/8Xdw/FC9f/RmtC8bVHVNi+PHj/Oo4vFamQOsc/RYrwJGKFQWb2Ud5RGIly5dYlDKdpV9yqNHj+KnR1+08naPd0uPHt27Q5dcatIPaHCLgctOxZ0zjmzfIsK87+LocOCdO3CrJi9vPEitwgJPibdAa/lNwaq8xA4FPNHZLA3ervKi691WDUWg+QSqkL9l7/7++cRojaB5ZPMUq1f7snJXT4DlmfXPe+3M7/D09DS7KfntNM1Cnj05gDAhy4uIVw6jETkePfpABLPEunfHG9rhw4fd61s1y4sOaE2TR1bsJZ3y7GJM/YyzZWs1RHVZJe6nWEOxudfPVBOhz+tTpibdqMzS4P084fHeDQiKojsECo2kdPl7b/naZ6trhfq8xdi2g0MDY8Nbiio84f14eHi4D7WvYYwCRv3kETTGjveRTzORRN7Nm9kwp4YoDwhjbQkURPBTTz1Fj1xyhEllx5rRaiikqampaHn+Epjfvn07v502WmBEvD8ltxQSD8M8T4OzZ8+iUwtnyFOaGAs32xaD1j3bFrflpwiUTaB0+bu5+GGpMQyOHyrVfopxnqrsndS4ap45c+bcuXMpHlZwif28xcXFPGteTifPnz/fsbUNKZaHCbtcZSx47I/iWOxsZwjcRTBG8kR3/vz5wqNj/qDj83jV9rbI1qJCAGYeUzTnuZrHgtWWeYvN2Hlr1ezq6cjIiF9oxQ6Enw9qJQLlEahA/pb7k2c7avrmA9td9Wrfubm5knbCss62JijgnKIqa8il1odnnm/s4dvp06cLfy9isqV/yhGIYAQHPiSl4eHhI0eOJF11KS82OrQRnrv02+E6TLmiooNnHlPoVJ6rRQkvpiLW0udtHm+73Zax6HaALY1ObhdFoHT5+8fFEuXvtoNDdf3cL/K3xqcD2jfnMlPUBDJ2WD7ZA84p2owpvyPrnF/DZrbKP7iTk5MzMzNFRYc/jhqRaghcuk4ZEZR0TseILr8RPHSPK6fDDW8OiqI85FGQ8znAcxXNyvMkp0toaOxI+3oPrvdXYnIOnJqLQDUEypW/95av3b+9UV4kNX7zobyg0i3v2bNnfn6eZTu9WvVXWfbYccy58lXvdjN7ZNlO3SJ18ppdUjbFvRc/0wfNcQZRa05djsgXuk4RwXiVf2U9e/YsjvEW6uJStI7RRpniihrpRgkMvb/7wV0fhcDLT7QwUwlT6OjRo3nsLCwsMD2kfcEOCo4eiVvYo5WaiEBbCJQrf8v+4u/A2FNtAV2In2hfNkVQD4VYK8OIFHBRVPOs/YEPFy5cGBkZ8V7/GE2WQD9thIIxIjhwJpwpJDocIzpModHDxtPzVKbJ6OiotJEBlWcfHf7GSPjInCnkNdjMH554YeM98+bF5tixY8zAnpU7XwEaeca3qXzklwgUQKBc+cvubwE+Jph4eM9gX+3+Gu0bu94kEKqnGM108uTJevruUK/sXeXfAIbH6uoqUgBrmUQwg0iTycnJnBoi6VWNjy8KiQ73jEhijWelJ96URAX6RZzRJKVa/1wCCKPMO5JfyClb+Lxg+Nm0WjF72QbmoceE5L3Fuho+5Sp1CGd0dJT3ovClvs0bINwjfgSA6ddQrUSgFQTKlb+l/uLvo/X93ln1Q9sW7WvIzM7Ozs3NmXxZxz6wy+pVVJQIAkQwyg+ZiA5eWVmJWkYM0SMCkU+0JycnaRKtk6mESctMSGqScimpSVI5C/wzE0WbAAAQAElEQVTZs2fRPQSI4EZ7kQiTRIbEQk5QVDh//vzt25X+wBmblw839T+A5BllVGnSiDAQp06dSrqatZx9eibk3r176ZH5yYAyeQDLdCXPlGZ8uUqdPOFk9crUx4GmDu/DAMkz2xlEE6OOItBJAiXK38/XNkr9xd+BvpG/hw8fRq/w6G/RFGSVkgLOOV4sPwVqCJxhL+3s2bPo4Keeeoo1G/voBhJ5EmKI9fL8+fN5lkx6CRLKANEZnFoZ5nPhnxIQIHuZbO6SCJNEhoQqKiooK4p+PmXmpISPKk3ZHk5pmHLp8uXLzE8GdHp6ml3hyclJ8mfPnmV8U1p181L5UaWPb/n9qwcRKJdAifL3T8vXSvV9cPyvSrXfEONoX/Y5UmREQ/yMuiEFHGWStQQFyQTI2sqxPmIR3UByrJ+p2pEjRxBA6U1KjS69a13NT4CN9hQjPLIWFhZSKuhSkwnw2OH1uMkeyjcRyEmgRPn7xzL/wYttB4e2DQ/lDL75zXkGpWjf5vuPAl5aWuJD8Oa72lgP0RCtA4jDuN0TKQqJz6+p3LOmKjSNAI+mnvKIDf5iP75oGoQO+8Oju8PRKTQRgECJ8rfU3d9++ObD8ePHW619mV4klkCikMQBhV9CZKAR/drW1Qrti7R16Z3pMTs761JTdRpFoOfWvvGWDf5XXnnF5As/ymB5BNK39svrV5ZFoDICJcrfz1bWygtj+8iB8ow3wTLa99y5c0NDQ01wJqcPSBwp4DwMWYrOnDmTx0KVbefm5jJ9a5B9psK/BFxlvH3Y18GDBxk1x8B5jrFV7FhZ1ZpAgNWHt+4meCIfRKA8AiXK33uXr5Xn98CXv/hbXg91Wubpw5pRpwdF920UsFZBb65stjErvJtX1hAn3YVR4BUbwDQMTpVpOAH2dN095B2et1/d++7E6q3JJ3WZxrdeb9W7CHgTKEv+3iv57946vPt76tSpjmlfMzulgA0H7yOzouEaEfdw0i9AGtLcr20ft6ohdIRs1jccFPDCwgK6qgZ31WVGArxpa+s3IzNVbyWBsuRvqd98GDgy3ErYDk4fOXKkw2/erILs8zlgUJV4Ak3WiIhX3Iv3262U5hhxq6ta9RBAwjJMHn2jqJaXl5HOHm3VpDICDFCHF6DKMKqjqgjk6qcs+Xtv+aNcfqU27vDWb2rcuigCDyE++HygaSDOnDmDY/m9wogUcH6M5Vng9ZWPcfzso4D1LQg/dNW04t2GTfpq+lIvIlA7gbLkb6m7v5K/tc8bOVAjAbZn5ubmWKtq9CHoGjfm5+f5wDQoyZlBARNdBiMlVNWPFcRCPXny5MTEROwlx0I+/0EB1/uGw4w9ePCgo8P9Uw0sDA2vKP0TsiLtcwKtlL/bhvf2+bAp/D4ngAphreKTyno54ACfaI+PjxfrxsTExNLSUl0aBXFWoJovlkyN1sDC1m9+B1DAvOHwcUF+U34W2OCUyIuiA4v3vn7UmkqqI6CefAmUJX9L/dO3gbHOfvfXdxzVru8IsFahgNmQqyvyU6dOoX1LUhJEh/Hqd2EReYizupA2tt/CsfCCwRsOr09VhswG58WLF8fGxqrstPl9gYWxEJbmj5Q8LJZAWfL3/u2NYh0NrG072IWfwg3CUUYEvAmwkcaG3MWLFyuWEUeOHPnwww9nZma8PXdpSHTsSKFXKtsGnpubk/aNDk1JWMwbDi9RyK9op4WXMIt4XZTIs8AaLIyFVa5TEeg8gVLkb6lbv9uGJX87Py0VYAYCrOhslKJRWMkyNPOqivBFjyIjStr0jTpFdCsrK3xWXqpI4v2BDbCJiYmoA/1cwoxiuEvFwksUs5fd5VI58zECvUjkWZA7g8WKS6ci4EKgFPl7f62srV9C0t+9AUFJBCwCaBRkYnkimJUSJYTwHRsbs7qu4JTPykuKDlXNBqS0UXQQT548CZYKhptXKTbd+TyBHhmOqCd5SnixYd7yMQIfJuSx07G2vNjMz88LS8eGVeFkIlCK/P28TPn78NBgpghVuT8IKMoHBIwIZiOT7bRClATqgZ1XpAkrZQVK6EEMCf9DvpjoUDOFRIcCIDRUNRuQCX32aTF4GfHZ2VmYV4YAEUyPDAevcLxr5e8XIyi8ahR8fm8rs8C0N+97hf/FamUhqCMRKIRAKfK31B/9fXTsqUIilxER6CoBPuRlO21tbQ0dzFJ35MiRTFKYBRLpgDREA6Ee2HlFmjSHFSrcRIcO9ogOGrSCDEqL0KpUeM1hGOsJg470vHXrFnjrGnGGg5cc3rVwA/HKfvDhw4djvY0tZJ4ThZm6GOknhRfL4+tCbmreakDKtOd9D85fX1NOBPqSQCnyty9JKmgRaBwBdDBL3eLiIlIYPYFeRN8g/mITSyMVqMYCiXRAGtalgRw5ooOD6FDqOJ8SHVeRvPfv34cGrSCT0guBxyJKL6RV1CaF6a1qvwo0Q4ZBR3o2RBjhBuKV/eDl5WVGDQ8ZQXRtLC7KucocYJ4TRc+pS5ixdtILo4NLCZMwvVXtVw0ZbmreakCKz0oiIAIQKEX+lvpvXvzZyAH8VoohoCIRSCYwNDTEUs3Cj/iLTSyNVKBaso3mXhkeHsb5lOi4mi55Hwr9h7VYROmFtArZ+DJLYXqr2q8CzZ3Ml1FV/n94yAiia2NxUc5VUDv6RcixdtILY43Tb3qr2q/ioTuZ2BhVKAKdJNA++fuIvvvbyZmooERABEQgDwG1FQEREAFnAqXIX+feVVEEREAEREAEREAEREAEKiXQMvk7cGQ4HY+uioAIiIAIiIAIiIAIiEAKgVLkb6n/7EVKMLokAiIgAv1MQLGLgAiIgAi4EChF/t6/XeI/e+ESleqIgAiIgAiIgAiIgAj0D4FMkZYifzN5oMoiIAIiIAIiIAIiIAIiUBmBlsnfgTF997eyuaGORKCNBOSzCIiACIiACPQg0DL52yMaXRYBERABERABERCBPiWgsF0JSP66klI9ERABERABERABERCBDhCQ/O3AICoEEdhKQGciIAIiIAIiIALJBCR/k9noigiIgAiIgAiIQLsIyFsRcCAg+esASVVEQAREQATKIbC5uOKR9Ovy5YyGrIpAvxCQ/O2Xke63OBWvCIhAKwjcODrnkT6ZersV0clJERCBZhKQ/G3muMgrERABERABEfAkoGYiIALpBCR/0/noqgiIgAiIgAiIgAiIQKcISP52aji3BqMzERABERABERABERABm0DL5O/m4oodgc5FQAREQAREwCagcxEQARFIJNAy+ZsYhy6IgAiIgAiIgAiIgAiIgAOBUuTvtoNDDl1XUkWdiIAIiIAIiIAIiIAIiECIQDnyd1jyN8RYWREQARGohYA6FQEREAERiCNQivyN66iYss1L+u5vMSRlRQREQAREQAREQAQ6SyA1sJbJ39RYdFEEREAEREAEREAEREAEehAoRf5uK/PLD5+trPWISZdFQARE4CsC+n8REAEREAERsAhI/lpAdCoCIiACIiACIiACXSCgGJIIlCJ/kzorpPzztY1C7MiICIiACIiACIiACIhAHxIoRf5uH3myPJT3lj8qz7gsi0AXCSgmERABERABERCBrwmUIn8fGRr8uoeic/e1+1s0UtkTAREQAREQgY4SUFgiEEOgFPlb6p++3Vu+FhOHikRABERABERABERABETAgUD75K9++cFhWFUlQkAFIiACIiACIiACIvAFgVLkL5Yf3lPW9x8+W9UPnwFYSQREQAREQATcCKiWCIjAVgJlyd/tIwe2dlTk2ebiSpHmZEsEuktgc3Pzo48+Wlr61S9+8f5bb/1rkCi5cuUKl6jQ3egVmQiIgAiIgAjEEChL/pb89V/9+EPMWPYsUoX+IbC+vo66XViY/5d/+R9vv/3W0tISp9dC/1GCIOYSFajGVZr0Dx9FKgIiIAIi0M8EWil/9fXffp6yij2dABu677777ptv/k/U7c2bN9Mrm6tUozJN2BumuSnUUQQ6RkDhiIAIiEBAoCz5u73cn/7Vjz8EI6iMCHxJAOWKfmVD97e/Xf2yKOP/sTtMc4xgKmNTVRcBERABERCB1hAoS/6W+uWHzUve3/1tzcDIURFwJ7C+vs6OL8oV/ereKqkmRjDFfrC+FpyESOUiIAIiIAKtJlCW/N1e5p++QVx//QYEJRGAwJUrVxYW5r13fLEQmzDLNvDNmzdir6qwnQTktQiIgAiIwAMCZclfbG8/XOKPP9xb/ogulESgnwmwO8umb3nbtDdv3nzrrbdWVz2/StHPQ6PYRUAEREAEGkZgizslyt8/K3MDWLu/W4ZRJ/1HAO3L7mzhm74WSHr5X//r3Q8++MAq16kIiIAIiIAItJdAifK31O8//FE//dveSSfPcxNAlaJ92Z3NbcnJwM9+9l53FLBTxKokAiIgAiLQZQKlyt8nyyN3//bGvWX9/kN5gGW5uQQq1r4GxC9+8b6+B2xQ6CgCIiACbSUgv78iUKL8HRgb/qqXUv5/c/HDUuzKqAg0m0CV+74BCTT3u+++yzEoUUYEREAEREAEWkqgRPkLkVL/+k1f/4WwUr8R+NnP3sv/nQc/aOvr6++9955fW7USAREQAREQgeYQKFf+lvrXbxsXrjaHozwRgQoIrK6u1vsd3N/+Fhf0QxAVDLW6EAERKIuA7IoABMqVv4+W/P2HjQUpYAZRqS8IbG5usvVbe6i/+MX7tfsgB0RABERABEQgD4Fy5e/A2FN5nOvZdmPh33rWUQURiCPQvrIrV/4PCrh2v9fX1+vdga6dgBwQAREQARFoO4Fy5e+24aGH9wyWx0i7v+WxleVGEUD4XrlypSEuLS39qiGeFO7G5uIK6e65pU9nLobTndmfU04qvMfKDN5bvob/0dAIkwcplz5bWavMGXVkCHy+tgF5EhOMgQgnSignmZpFH5tijwBJZlquz1w0iVMKWzEh8ZPEYIXHjjwllJOaAjrih5l7oMbbJqTqh7tc+QvwUr//cP/2g2cHvSiJQLcJrK6uooAbEiMbwPjTEGdyuoEoZKG6Nf7Gx8NnPnr41I2jc6S1yYX104vh9Mn0O5STqHNt6IfUZ8Ggbc7eS23OckJoaxPz10dex+3ro6/jfzQ0wrx17A0uffzUFwTG5j6ZeptFkdWxVPf61jiSiMlzY2yOifS7vT+EPIkJxkCEEyWUkxg7alKfVrRtOzdCIBDuoOgd9+npRZOYpQRuTUjmcxNi566P+s9ghceOPCWEQGL4iJR4aUXb2kPgXXdtYh6XzNwDNd42IVU/vqXL38HxQ6WO9x/OLZVqv8vGFVt7CHzwwf9tlLO//W27/wCONZg1AFWBKGSh2rhw9bNV171P3rqpv356kbZYwA7WmjM6LLHoV5Y31AOh3T2/fO9yhp9I37y0cufs+yyKrI7XR15HQEsHFzK4RnYwYZBETB44M5EcLVOT+rSiLRbWJuax5ti2IdV4oUIC4jwhEMiG8x1H4GZCMp+ZzDqf7QAAEABJREFU1cxtZnj1QQEc7PjPXb9+enHD2X9c5dlC/fVanxjcxehv/Oddl2cCLuFYn6fS5W/5X//VX7/1+Rzufvjs+167lkHBxBLZtWvXgQMHRkdHv/GNg+Rj67gXrq6uulduTk3WAPQciyhrMGsAqiKnb1jADtawiWXs5zSYpzkKgz1Cluc7Z98vZHlDNyOg0cEs/I2S+Hkofd22khxTIiw7mDA5u8UCUw4Rw5TDMvZzGiy1OVt6TB5UFy9UGxeu4nye7pjVzG1mOLFXc7uBl47oDuBgz+k/sWMBOxU/MZgnhLB+epHe8UHJEChd/m4bHtp2cMh0VsaR4eS1rAzLsikCDSFw48aNPJ6gel966eVXX/1PL7/8N6Ojf/3CCy+Qp2Tfvn3eZlHkH330kXfz6huyjJk1AD3HIlq4A9jEMmsMvdBX4fbTDSJ86RqFwVZZek2/q2bBRltLBLsDZBqwVcnLQ0mygymHZca9linXk4MRvmzZMnlYpnvWz1SB2Mu+3Rg+wIKXjuguk3sulbGJZezTC325NPGow0759ZHXmSeFD4GHM01rUrr8JeCyv/+Q5/cfcE9JBBpOIM+/c/H000+jep988kkrRkrGx49x1Sp3OWXzmC1kl5oNqcMbcjVrAGsMKw1LGjtG1cSOHiU0hC+radk9oq3ZtVqbmEfZlN1X2+2jaR5Mg7Ol/0pg9VOu59AwPZgkRvj2rJynQhA7tPPYibY175Pcy3QRvVpgCfbp5frI6/RYoFljCpvslPMZjjnV0SJQhfwt+/sPvFyW9/Jk8dKpCFRPYHPzj36dsr/73e8+n9KWq+wNp1QwlwYGBqg2Ojr6ve+98Hd/95/ZPGYLGQFtrjb5yJPh1vgbfHBZgToMOLCksa/DksbWS1BYeIbQ2FxEj+ZZ3jy84nn7+5EH3wn2aNsPTRh0hh5NwzSoLF76qmDKuYTDix/Tg0niUrmQOsQObZhDPr9BtDufcvA+idn81hwt8HSiR/qld8cmPavxSoDNntX6uUIV8rfs3V/Gj7ccjkoi0EkCm5ubfnE988x/7Nnw+eePxNZBOj/zzDPo4/HxcSQvW8ijo3998OBBpHBs/QYWsunLDtxGTf88JKqUrRcWoTLIsNLfHJu7U/7mYqzzKAPEFu8VSPDYCn1biPhj0Bn6WgjQLxKqrtUQ6UbvTAymR/XhEzvkc95uPDHQ7nzKUb3/9Ei/9F7I8K1NzK+fXsSmUgqBRx56KOVqYZcGXyn39x/uzJb+GVNhLGRIBDIS8P7ur8t3G4JvMpCh/ne+8+xLL73893//X8bHj5GnZN++/Rn9bUR1FkI2fWtZicPxswihCYqViSyQ2GS9D3dUfX7jwtXrI68jxKvvuoE9MsRrE/OIv3p9Y8Kz54cn+FOlJ/UKxyDS9dOL3Bp+sfNZSu1PjEKGj0dflbvvAfzWZarY/QVK2RvAfHawubhCR0oi0D0CZX/N4Nlnn2V/99VX/xN7vez4lt1d2QPE4sfyz0Lo1lHptdjUuV6cTGRtQ9+wTJbut0MHPHhRG0gfh7pdrsKUYzO+OZoDT/AHr6qBzpysXTgGkXK7ETtb0UFJzwygmMZ1fZYSdS/P8PFu3JxHXzS0RpVUJH8fHf+rssP+wzn9AHDZjGW/ZQRu3nT6yQj2fVv0lYb0MWAlY/Fj/UivVvFVIxPzb5SuNe8zTYQ40odFt2KkzemOYWXK1b4ZbwHBH7zidrDKCz9t4Jwk9t87v3CCCFCI5sLJ5DFICNedQwg6QvTfnnonOHXL9G+tiuTvI0ODZX//gQWPse/fkVTk3SXg/fWDK1eudJdKTGRmJWPliLlWdxEykR0mpJK3I2sT8zzlvJuX2pAN6f5UwAzog2HN8g+LlDoQYePcCAg7bopwYYF5LBN7M+ek4+1GCCACVIFYijLl8c7MI4LAi3Kg83Yqkr9wLPv7D3ShDWAgKHWPwGOP7fILatXxn0r2s96wVk1eyQwqVibkAoLJnGY6srA1U2cEUfShAmbKMaAMawChaRmEHfIOPwt3DJtYbtqmaThMxiX9rzNNCCAKt2pUvmcIYW95/2zycIRdbUi+Ovlbwfcf7sy+z4RuCFm5IQJFEWD31+/LCZubm1eu/J+i3Gi4nU+m3m7ySmbosZ4hmLIq4E9nLjZc+5roUMD98zcYrDXoPwbUxN7YIzcFfuJtsR6iLLFcrM3CrbGBmhL71yEU3nFxBtNDCPphfPW1h4CGY6Y6+VvB9x94Et2Z/blj5KomAi0icOCA/e9WODp/5coVRLBj5fZWQ/u2QiBCmMfU7Yl5livyLolNnRb9LcvN8TeyinsXCA2s0wrxZLihU5lyJl/Ikc8i2rLRSOw8HKJRtyuEnsP3x4V/48ESDVMlKQSqk784UcH3H7QBDGclDwINb3Lw4EE/D9G+V7q+AbyxcPVOTb+A6zcoLMk91zNjGSnZrk0d1mBCcxf3JszWHdmPb4v+M2w3LlzFZ5PPecROW141TaR4yzukyZsjpxSafCuODF/61t6nM4utCKRRTlYqf3dMjD68Z7DU+Hn4ps+SUnuXcREoiYC3/MWfK53eAP5sZY2NHMJsV2I9Yw1O9xkRiZTkmZZerWlXEffrMxeb5lWB/mwurqy38N8UwGc8T+LgWM6rJnYcKzen2p9W1gJnynulHDgybFIZOueT6XfwPIginKH8s9WvAwxfUj6FQKXyFz+0AQwEJRHISmBgYOAb39AGcAw2tG/rBKIJg21dBK7Jxx4RkUjJ2EsNL2QzHp3UcCf93GPImHJ+bWtvhef47+1GG181tx0c2n9x8rGZo0HURb1SonEHXzm0+8yL2H/y/mnS/sVJkw6s/YDTx5dODM2NUyfoOmcGz2MtbCz8W2y5CtMJVC1/d049l+5Q/qushdoA9sKoRo0moA3g6PCwgVrsZ9AsljtPPrt3/jWWLhawcKKEcq5uP3wg6olHCU+qpPUMa2zUISLJFJhYsNmd2nVqLJyIiMICezGmPpl6O4/SMkYaeOSdpPCdNmbUjuMj4UEhTwnlxRLAc/z3tol6ZtJ6N6++IRifWJkeGBsOuv505mL+V0qeEuhaLO9deA1JE7YfdERm+8gBPvGmzl/e+gGecPdRmCfheay24VmRxyyOodHxsN60bXgoTxQebauWv0yIwm/paNj6BnCUiUraTuDpp59mD9gvis3NzV/8omv/MDjqig1UPyDRVqgNdnFY0nbPvsSHVDyprDqUUM7Vx5dPPPHhNPWtCh6nGxeuJi1drNMeBmObsLyhcZHvbEqxO8VOWDgREYUIfcR9IUEZH1BasUu1udrIY2+n+JS5wHcSlsIHKurDaWbU0Llj4UEhTwnlzDT2F9FbvZ1zq4H/bOK61d1Si9Es9lUzsA4HJp7RXkxCksmjybwDxyYTHoxBL2QIHG1AxjtxKz0YspVpdO0jQ67f5KQmnvBs4Tb07to0/HRmkeeeyQdHpmWQz5oBNY8FNDoe1pu6L38Zm51Tz3IsNfGGyr1aahcyLgLVE3jmmWe8O/3ggw8++ugj7+YNbMg9zp2e3zH2Plkph84dS9rFiXbBk5r6SBNW6OjVTCV/iPvnKova1Wa1Znlj3UXjIt/THUPcm6DQIuk1Ha+un15EcDhWbkU1trQL8RNxxrsW6hYVxVxKscnVnVPPMYKIQm8taNlnE9cq6XmK5EJ49ayWqQL3DkGxMwoHJp7RXkxCksmjyQicuyzrCwBzHpvRCc8rZZ4nBg7jD0OWKcygMiKY25CQuSuDwqwZ/Lf27xkaCrPaMfUBBWqT78Nj1bu/IH50/K/yDD8WXJLnk9fFtOqIQE0EnnnmP3pvAOPyz372HtvAZDqQeOjfmS1gP5sFgL3P6ErpgghpwgrNbpDfAw01Q1sW/mhfhUgNZP1fLJ9geWPdjXaRVEJQuIQ48wvKMovgsErae8o+fSHbn0w5xJn7u5YhhihEe9HWnOY5EgWxZLKA7vfWWFZHzCuiQNRy7xBUz8nJhHR/AWDO8yrLnLc65ZQ3sTy/9oDPONzTWzpKT4TMAwcI6dVSrlr7939avpZSOeUSPgA2pULnL9Ugf5lAfz4xUgFZ7tgKelEXIlAZAbRvng3g9fX1paVfVeZtqR3l3/rl6Y/Ii10pM3nOblDW9cwIX9QMbaN9sfX7We6/42YHF6+QDlH7aSVfXUOc4R6blF8VeP4/goMXFc/GDWuWX8rnn3JM15zbhwbqH+I+czCXosecwjFskA1U80rmMTMRjsxJdoLBGLZp8hRyiTmf9CqbZ/h4TYW86Sj/EQ/xE4e9TfH0824bNMQNxFhw2oeZGuQvlKt559hI/l4dPiiJQBsJ5NwAvnLlyurqahsDt3y+k2/rl7WHFQiRZ5n1O2UhwRo2ezanDttIbP7FCl/T/A/nlk3G+8hqzQ6ud3PTkKVx3+IkesWceh8LWaq9ey+qIRKQTdM81hh6Jkn+KYcKxA7W8jjDawkROVrIIxyDLnCYackGqofwDYyQQTwQvvVixqYvqppLVIhNvIN5/xQJbqfcrbHd9SzkibFn9sWe1ZIq8IggoqSrjuX5p6JjR42tFiN/K/CVGyD/U9XFT48vObmYVR0RqIsAG8Df/e7zeXr/2c/eYxs4j4Xa27I/muejWFbiByvoSDE/4GBosJ7tW3jN5GOPdIrwZfuKbSSUZWwdChElOWUWvRS1WuPnnnPHLKmBk5nS3dxqPlN3JVXOKeIZ/QKnHJMNa9jME+wf3DaAmZBo5Twd0RZXcbioaUn4vJiZaYlls+mLqKCjpOT9j6LxKUpRblu+bRvei/NWoeMpTz8icqysakkE6pG/eJPyosbVohKfIRby5lqUP7IjAjkJbG5u/va3ubZvsfDuuz/lmNOTGpt7b+QYn9l3YQU1+ejRu4TdFKRnbHPKewpf0zCnzGK1Rl4bU4UcUcCB1PAzyEM4z1+m+3VaeKucU45Xo2KnHNZybvA7vpY4quQU4Ig8tC8Op9TJeslMy50nn03f9A3M3pn1+TuBbQeHds++FBgpKrO5uHJjbO7G0TlUrLdNv4jC3Xl/aThspNX52uQvSwUfWFTAbr1zf31cATR10UwCaNa33vrXDz74IKd7N2/ebO/voPGp38aFq94E+NyppO0cXEJ6mk0p8iahR5/4cJpyFmxTkn7MI7PouozVGs/ZA053O/0qu/XpFRp+FfmOiPd2kpcf1jvv5kkNB8cPof+SrvYsJyLi6lnNUSUn2SlD+5q+mJbM9m3DvX8slg3se5d9/j4M+/RiuivkGAjfnJ/w4AwRmeF72Pn312gVTqZ5uMTk++dYm/wFcTV/AEdH+goEEJTaTsBoX5RrIYGgoa9cuVKIqYqNsIR498hinFPJ9eyaJdPUMcKXLTqXFdo0YZ1GlJi8x5Gui12tAx/YukPDBadZM3k0fda+yqi/keNf1eKdhJefMrzCJiPODiUZv7S5+GF6QwYuz4Tkdit83zfd4dirPcOMbR/87DAAABAASURBVMX2HC8YsZc8CnlqmR3f/MI36N28VXJvBiWZMowsXmVq0rHKdcpf9mDy3LruI8GEy/mRontfqikCZRAoVvsaD9kARgSb/NZjo8/8FjMT0mMzYyUJRGOfI/t8Q3Pjjy+dyCR8aUja7CVHqJOUUNt0nXQ1f/nOqee8H9cstCj7/D7UZSGPSkChluo208zbPuo2vW0e3Y9lbjdvcUbzotIfF1c8TDHhPVpFmzB5Che+ppdg+LxvTHYGW31jGg7exzrlL05ze3CsIH0607UfYK8AmrpoCIEytK8JDQV88+YNk2/LkeXEz1X2oopa0tId4MXeb9X3W6eNM4/NHDWZko68Njw2M+ZtvNWftLKB4hc4O4ilvpPgFfbZYCbjkXrG5X2v4QyxV3O70Vd6yh7FQwjK/Fu/9FuS8DXxBm+VzAFTkvWIhd+PvP7J1Nu4Wl76fG0jq2PV1K9Z/rJOMM8qCPX+7Q1edCroSF2IQLEEytO++PmF8bdcFPD6+vpHH33EbvHS0q8cE/XpovB0z+trfLhR2bet6Msvef8xyuArh9y/YuHnG63y/ItF95bb+i8OIguI3S9VM+V25viHVFNeS7iEPPILnFZ5tqVpXlRCe3lEkVP7MmdKFb4GTqCdHh0bNiUeR6TRnbPv3zj64E/xSjr+bu8PP3r41PWR19cm5u+eW2JEPPwso0nN8peQ8uwo0Nw98aarX4Fwx9XvNRsT/xcbtDfLc8dRAX/wwf99++23fvaz95ac/yvqa8rh2FmSw6eZ8rxsZ6pffWVvZZ9ztXaMlA1g744QBI69NK3aZyu3/Fzi04ZqphyvJX4e0irls+8/5Xhj2XF8pIL3Mfzvmf7k9Y+iDY7/VU/LsRWQdx8Pn0FHojdiKxRSiPAdmht/YmXaQGYCMNkKsVyeER5ud88vr00uoIZvjb+x6fWNlGLdq1/+8oBgLIuNKsna+unFe143Q5JBlYtAqQSWln71Qe7feejpoaMC7mnHqrBv3z6rJP/pfd/P0XjI+H0hIb/PjhZShEhPC6x/PesUUsF7nylPdIV47m3kTytrfm29WWXtjteSgSOe+3/3kjTuQw81+as47og+83p78fg6gRG+yDuPzWb3cBjo/RcnEb4Ip6AVE2Bw/FBw2vzMxoWrvCGwQb5ZqwiuX/4yVFV+SsJrR3P23oldSQSSCKyuri4tLSVdLbbcKOD19fUks1RIulRl+R99/zjMYz2rMi768haI2w8fYP3DQgVpYOwpv15K1QR+Ljm2+pPvjslAjo+kHX0Lqnn3lfI+6T0hq/kqThB7esbj7YUbKt2mdbVK4bt/cTJ2rB8r+av/VsiFnLJBjghem5ivS5I1Qv4ynLzTFAK0pxGewrcn5ntWU4WHHnpIEGokgNz82c/eq9IBekz55zBu3GjZX8hZ6MynhFZhN06rDI2+mv8xa7HD6r02bx95slhPUqw96vtakvJxKOokpceUS43aiUzR90khuP+SbhOEr4mCG3PH8RGTb9fx7vnl6yOvp8zD8sJphPwlvCrfXdh415eAYa7UZALvvfceerRiD2/evPnWW/9afb8VhOmtDyrwzXThvbH9Z4X+683GmZTjdt/ualnhUgLpeSlnBRRJTgs1NvcW/fhc2Vdx6Ktn8ph1LjO8OcI3IICIaumrKZuSN8bmQBrEUk2mKfKXDWA+MakmZnpZP724seD/D0dhQUkEyiPw0Ucf/fa3q972BwYGvNs2XAF7fxLtDUQNiyLgsQ+X0rX3R/MpNou9VKX8ZQEt1nnvG217hV/FKTbkwFr67i8q7ePhM9V8xzfpqw6Bq+EM8+2xHD9NGDZVff7+7Q2QwrbKrpsif4m57J8Hp4twWpuYd3gvDLdQXgQqIrC09Cvvnvbt2zc+fuzAgQPeFopSwI8+6q/Ck5zPsymVZFPlbSTAjlEb3e68z4iwrsaIOGum8A2A75x6rsptxKDfojIVK+AGyV9umzz/rmbWAeBt43Z937nO6q3q9w8Btn6vXbvmF+/TTz+N9t21a9fzzx/JuQf87rs/Dfvg8Y2Iffv2hy0o30ICclkEshGo+Ks42Zxzq219sMAr96czFxsufIPI9pw7xgZ8cNq6zO2pdyrbl2yQ/GWceHfZdnCITDXp3uVrt8bfqKYv9SICjgT+3//7wLGmVe0733n2u9993hSigIO8Kcl6RIKH//aOLeGsFsqoz0tyGWZls0YC3n/3bMmUGkOovevCUdxL/kG02oMt24EAZiB8108vlvppw47jI098OJ3pqw5JEB4ZGty3ONleBcy+5I2xOcgnBVhguYP8LbC3XqYYuYq/ArF5aWVNPwTRa1x0vTICbLL6/dDv6OjoM888E/bz4MGDVkn4qkseT8IK2KVJUCfP3nNgJJrxlr+fef38Z9SB8kp4+pVnvEDL3l/hLXxfMJApBUZXoKnKNrHwuXAU3vqjadPYwx9gEr7Z8UX4IsggXFIywnfo3DHvJ1vUMUJuuwKuZl+yWfKXgRwcP+S9GUBzj3T3/PInU29nbTg2Nna/hP8WFxezetKu+uKWPl4ffeTzb8MeOHBgdPSvo5bZD96X79+eMArYw6uc/UZjyVni8fOfX/ZY1f9t9/2drM3Flap8fNAPH5o9+L/s/2NVjm2UVB5bOVzY8B1KVFTY21Lzn/m+3XnDTwrHexonGcxZ7vHSxUbvx8Nn2ih8A1YMa6sVMPuSd8+V/pv3jZO/jB9vQhX/fseds+9XwJrQlEQgncC1az7y9/nnjySZffnlv9m1a1fSVZdyFPAvfvG+S80K6vBY9+tls1qN6OekX6sqdxnL6MtDoBhQZThjLIeP2/1/6M3nXg537Z73RuEN3923NtZs3Y5vFDKPyseXT+w8+Wz0UraSmmrfnnqHPfhSO2+i/OVTgOp/v6PiPzksdVBlvL0EPP51iW9842CKwB0YGHjhhRc45mHi8cXfJ58s5Tf/t3tvkV6qdIvUg7a3EGGp9lY/Wf3cXPwwaxNTP+UzvT8b9vx7j2peadJ/BstEF3v8Y4W/rVkNitgwG174qO8/CFJGXGV81SHdz92zL+2df63KP6lK98f9Ko+1O7M/d6/vUbOJ8pcwdk49V/13t6WAIa9UL4Fr2X/z4eDBg+k+79u3/+WXX86jgNPtx14dGHg0tjxnIe/G3hYa/lPf7NZ4h7bpq0qz9nj33HLWJqZ+SnTbhveaOlmPfEhdwRcMvN+47l2+VoF7QKMX+iLjkR5tkjr08L9nE++3l56WM1WoXvgG7g2OH3p8+cSuU2MVf6geOOCduTNb7qeODZW/8Npz7hjHipMUcMXAq++OpaL6Tkvtcf/+fT3tGwXcs1qBFfbl+85xkifbhoe8H+IbC/+WZLYh5SlbpOkelr1OmN65d7xlVsredp5/suEP5X9BcLvvlx+AVoF79JJnk4wbCgsdTnmGL8CSJ1Oj8A3c5uXzsZmjB9Z+MDQ33qIfBmYDuNQ9i+bKX2Yt7yvB+FWWkQKuDHX1Ha1NzP9+5PXKPinOGuD6+nrWJtTf5/bzulTL+VNo9OWeSvryAw7wZODoke6eX0bAeTSsrIl3aOyDVvDXC5/OXPRGkb7L6P1Zn/dutHsgCETvN64K3Pt8beMP5zy35PlMnOjcUbS0pvdbZZ54mTMImCc+nB4q9Fcd8rhE2x0To3sXXvvLWz/YO/8a7kGGhKtcamYqdc+iufKXweB9hbEhU3GSAq4YeDXdoX0RQLxQ3hiby66Aq/Dx008/LbWbp59+uhoFvK+crV8DJ89mYR4BZ3p3OV4feZ2O0CUulcN1BnJ8Ev3pzGLYVOF53hy4fbzNpuz+YjP9KhWSUjW6/9Gx4SQH0stxj5mQXifnVbZ+eab5GRnwjcuvu7paVRwmahJl+cTKNAKmmW8XbAYPjh/Cvf2LkyR2hZ+8f7qk9PjSiTwqrtQvtTda/nK38ObEZCJTcZICrhh42d2tTcwHizerRWMVcNkcUMDf+17ev4Tr6WR5W790PTj+Vxz9EnOg1OcpXrELe+/ytfXTix8Pn0H6ZBLBedZplBZKCAdKStxB3pZZ/1hxU5p760tslq376SLPuNyZfZ83B4yUkbCMfW/LWbB7d1J/wzxPjEzeo1UC4Zs+4TOZbXVlPtFCYe84PuIXBY+1TI/QTL00Xf7y8rRn9sVMIRVVGQXM6lWUNdmpiwA3z63xN9A9YQeaqYD9VGPWH+U9ePBg2X8Jd+BAKT/7YEaQ5ynLjMl7HJFxTAmPhi5NsHx76h1TkzlmRLC7KmXJzPPNPIRgSR9rEMJmjp/OGBw/ZJgkHR/N8UrDAln2g3owh3tMA6ZcUuA5y3myYd/bSB7s3p1W35AnxraDnr8u4ugtT6RChO8nU2/z/uzYaYuq7Z59CUR+Dv9p+Zpfw56tmi5/CWDHxGieJQEL3onVq7wnl7dXtTVsYcfIkZtjcxsXrkZ9Z9noxh7wb3+7Go0uvWRfmb8FMTAwgMJOdyDn1cFecirFPmrpdmn/0GNUjjDNPpl+5+PhM46rWp7Q6IvQmPMp4XtcQlITgkfDoMlgL/mI7t9++EBQP2tm/fQiTmZt5V6fXZg87vHmgKxx786xJmsTnzM4Vo5WY1UFe7S8kyWDOZ4Y6UBQdYUIX3phQO+cfZ99Nx4jhd/F2K8xMdPKGwLvuFogf4ltz7ljZb+90UtsYtewe3MxNtLuFbIion1TVgjkQtMUMNox60CsrmaWv3RhFPC+Er6kW7b2xfmdU89x9E4bF66y0ng3T2rIHiRCJ/YqmptVzUUE59yQY7Yz5wtcO7mJuEdig3IsHDgyjHzsWXnHhOfHo8Zy2U/pnVPPPmR68jreKfpfVuJtirXJy5cvGzVQjnzpWQn/t2NitHCrBQpffOOJFAwoDyieFWV/TYtOq0wuD4Eq/aGvdshfXh32LryGu7Uk5iIrymcra7X0rk79CPDsYNlGDaQ3b5oC9tCj6+vrH3zwQXqYsVf3PdgD/pt9RSvgf//vn47trsBCPs3MsxuHJ6w0rDdkikrIkfXTPf74LBDBKb/mw7PO+3tyJhbmPM+rQhSw0b7cI8ay3/HP3XRtToEC26KiZiijT3teS5A7fgRMK95/sGzyOY+8aGEtjxFiyQk8T+/Vt+WJwWtYUf1Cr6gdX1xisrFU8UQiHyRuuhtH5/jQoJAbOTCrTJhAO+QvHjN9h+bGydSSWFF+/8UPZtXSuzrNSoBlhmcHTxCXhlTj6cNK71K57Dr79+/36OIXv3h/c3PToyGbzS+//DdPP12YYMXgo48OeHiStUnO3Ti6Y71BAReyumSSIwi1W8feYMrxhoYb0bQz3942BnleoQVzTukHN9HYHHcHBr0Tn9o5yqxHhgbz6/7rOZ7SzATG8drQD1GWf4j8nDDu5d8uxfJavi/e4CQWer5o9RyvnVPP9qzTsQo7c99WAGE+7z7zovlVB6YEJTnTndmfIy2SPjXiQ4P8N3JOD4tqnvS4K8ryTpHOAAAQAElEQVS+h53WyF9i4zGa8/mIEe/EMnB99HWWBG8LalgNAZYHlplMfTG4yJGcciFTj0mV97n9iK/VHO373nvvWYWOpwjW7373+QMH/L95Ge4ITxYWFrzleNhUep6nAUtRep2eV1HAOVcX5Agfu3vIERY83tCYddFVgVf9/DtVKGCMs7j2hBCtYILiJuK+sK5mPX1sZsy9yZ/n/oSaVwuPqBkCHhq/2/tDxtGEfGc25p+bemzmqHssSTWZctdHXqfHpAop5bSiLRZS6rhcYvOyEC3o0ldz6vD2kv+JQTjbhvcWInzREh8Pn/lk+h0z5bAcm7iRER68mMVebUshaytPvKZ52yb5C7vdsy/l/NATI3kSS8LaxDzLQx4jalsSAT5Fuj7yut/ywDOIhZO7tCTfHM16f3H2t79d/dnPPBXw0tKvrl275uihS7UrV668+eb/9PtKhot9U+exLNLKNIkeg9XF46Y2C9hG3B9WRjuKLWFJ2Ij7t+h40MXWz1R4//YGiytLLH46NgQCCy1N8gQV9IXa4C0lOO2ZGRgbzq/73aPmccHrAcHyHmI9NDAShbZteGjnyQI2TZly9Lg2Me/+tEH48nSiFfq+J8aeFXZOPVuIgOvZUdMqDOX+p2Thn/7RTc+QmXXmFkNLYK1nfVNh/fQiE4C25rRdR+Y5ewTePnPfebdNb+ghf9MNlnuVm3bvwmu8vJbbTap1HpQ5d4xSzeuiJ4GNhat8isTS4tn+oYdY83jEcK96W8jfkL3Yfb5fxkVuvvvuu+y/Orqxvr6O8EWnLi0tOTZxr4YbyPGFhfmsv8vm3gXSKr9aMt2xuiCDHBUJi1CwgDFnjAW/I4+yXXF7imwAF/VJF0ssC+2Dz/Qn5rlHcD7qKnMetccSFd4BjVbLWuIh4h+Lo5G1X+oHURMUg4V8DBKS95Optxnuj596sPdGTepHU+w/pcZgMWTRyh4lrCPs6l3/4l9IwbdYC5TjPHUQvrwpxdbJWoj/O4v4GkDWfptQv5D3KwJhLBgRphBziXuHkp6JalRmKJl1PG2SZl2KHTplgeM+ja3DVGlg4oGzNjHPquoRrwmT6Sr5a1A8OMJi/+Lkg1x9/0NjMaLM5vpcUM9bCLCe8VKeU4tgEQuMLI8q8nWlp5/+D95dsweMnP3FL96/efNGkhH06NLSrxCm1FxaWkIEJ9XMX37z5s23336LjpDmCOL8Bi0LHgLLshCcMvRGkaAUjWZipQlWFPIIEco/Hj7jvYAFfQWZx2bGeKUPTsMZQuPRHy7JkzfRcY/g/EcPn7o+8jrznMQS/uB09HUk8kaObeyobwNHhvm4OVqeXlKUQDG9EDVBrbNzdnQOvWISO+J3zr7fcz1GbURfFRisYn+HnqXEuMcoMPEYEZPIU4LDXKWOCaeQIzugRFGIqTYaIfyi3GYKMZd4h+Em4snA84GnRPDEIEMJiQFlNKlG5ZxDyXzmPqUvPqWxomCqNDDxwOGhituWt+6nbAS4V85as2W7vyY8iNT4Z3DGB0aU2Rw7EU0FHashgFS9PvI661lR3TGyPLAwW5TBrHaefvpp9oCztgrqozKvXLmysLDw3//7//fWW/9KYkt4aelXZJC8FKJHl5aWEKZBk7IzKGx2gv/lX/6H8QT9naLOMznDo2DXqQzfLnUxzgTYuHB1/fQiK02wopCnZOPCVdY8FyMudRCIKftwaJQCl2rLH5Zh5B2pwHDCXSDcvZ1H94dN1Zi/M/vzaO8FfuZgGWfiMSImkbeuFnI6+Mohj3eSQrpuiBG2zwp/YnAT8WRY3/rE4NFBCYkBLXY0Ny5cRXAjr2tCWmm3pU7XVspf8PMMKuRrWJjKk8xEZIc/jxG19SbAuzVv1azl3hZiG/K0qlEBo329vwFshXPti//YEkbvkq1S8lqemFPjCfr7/fdj/rTI1Ml6fGzmaL1/D5DVYVPfRSDy6C/qKxCm08qObJGiM/y645WmCc92nI/9/gPltX8BDx88ElNuT+4vv3r027QmLX1ihDGyQsX+zUC4Tjfyg73+xZw8YbZV/hIzmwRNWBuYiOzwaxuYEakysTvLpu/66R6/tOrtEsNaowIeHf1rb8+raNiwPtooRxwFIk+51ol7HstsT+SZI7tmjm47OJTHQiFteQjwcXbUFBvz+xZq+x36qD+OJfiM546Vu12tjU+M8IjwTODJEC7pZJ7Px7zfol2AtFj+Eh6frzEPyNSeNr74PCL2WVm7bx1z4PO1jZI2fS1QLH51KeBdu3aNjhb/zxRZAXbmlEckj4IWheMuENEr7Nixb9eW6Hgg5x8LokagNCHkpA3ggbHh3WdebIKHjj4MzY3js2PlzlfjiZHrL4hqBcSb4b66//ypGgB/7vYv5ng70275S9jMAx64ZGpPqKW1yQUEU/QPJmr3rTMObC6ulLrpa4FiTBlQdpqt8gpOn3nmPyKCK+ioG10Mjh9igW9FLGxpZBKI20cOtGWp5lHMA7mQUSDqJujLzUsrSc/znVPP8RpTSLBlG8HPnPvxZXtYvX0mWFueGGE4vAnzZsj7Ybiwk3lUftmTtvXyl3nAA5fHbkNmAI/Lj58688nU22xSNsSlbrjBInRr/I0bR+c+W12rMqIkBVy2DwMDA9/97vNl99Il+zwrmyCY0pHypGL1Sq8TvdqKpZqFmY1qHshR//1KGqIvY/8AzkTEawzK0uQbe8RD/GysezU6xhOjXQqYW4w3YZ4GNUKrrOsKJm3r5S+DwQOXFYWZQb4h6c7Z9z8ePqPvQhQyHLxIfDpzkZeKjUJ/mMndt7oU8JNPPqmvQLgPEzURTA35qymciSa0L+/qPK+il3qWNHyp5vFbxsLMEshmeU84pVZI+v6D6RQP0Zcm38AjvuFhGY51w2bDb6swZJ4eZdxi4S6akx985VAF39XpgvxlzMxXeXgEk29IQjOtTS5IBOccjgfCd/jMeml/4uboHqNZy7cgRkf/+hvfOOjopKpBYPfsS83c0UHGeWtf4iKZpbpRTzm8IrEw/8XyiZI2pdjawD691JW499M3MtCXqMy63EvpF6/wLaWCLkGA22rvfM3/lhZupCfz9CjpFkvvuvqr2w4O8TlSBf12RP5CipnBu1HT1gY+qUcEo5w2F1dwslup3GhYcnh5QPiy/JTbk5t13GAcq/8e8PPPP+/978C5RZatVvO/kcx61jQFjBDh6fTI0GA21pHahIadRj3l0KbIejYgIs4WUwA07NNLMea8rKRvAGMSlVn4r8liNk/afeZFvMpjoX/aDo4f4rZCdTUzZKYW7nEjNNO9Yr3i4cYbbzXBdkf+MgbNVMA4tnlp5cbROcSTfiEYGj2TEb68NvDy0LNylRVQwJ9MvV1lj/Q1MDDw8st/0xAFjBvj48e+851n8QrfGpuQiY8vnWjCesbTHC1eoBDhKfdgq/XwgSbAR9Y/vnyi7LUK+/RCX3WFzAP8s5XonxxsceexmaMMNMO9pbSOE3xgO3Pn1HN1dN7WPrmtmGN85t6oAHjr239xkqnVKK9KdYZHJWNRaheB8U7JX6ICHO9J3P/km5Z4ht469gY7msi7pvnWBH/Md3yvDf2wgcLX8OFhxIupyVd5RGuigA8cqFnxoH1xA2eeeeYZRHDt/qQPAY+C2tczJgyPI7R4uqtZr7LVSmj1fsuZZyxqj7Uqq/Pe9emLHU3v5jkbpvwBXGCZgWa4GfSgpPoMvfN2xHZm9V23vUfesni88+bA3K49Fnxg05fbvIKvwNYebOAAj5Qqp27X5C8cWfZ4BjF7yJeesnfAjibyDhH86cxFBF92Ax1swc7K2sQ8TJrzVYcoZdYVPoTlERm9VEEJohPpie6soK/YLugayYsb5uquXbvw53vfe4GMKWngkcGqaz3j+WNWLx5HJZHZPfsSO0O17HAPHBlGY6H2SgotySw7mrVs6jOaA2NPJXkVLme4kSwMfbiwmjxO0i+983ZUTY+d7AX59cTKdL3vlnzQgQ99tenL7OVpVvEjpYPyl3uSZxAKGL1CvpkJEYzU+93eHyL7+vlrwWyE3xib+/ipM3fPL9+/vdHMwcIrPhSrUfvigEnf+c6zKM5Ag5rCso90R6d0He3o4MGDaOLR0VHqRK82pMSsZ6wolflDX3+xfKKC1WtgbJhlEtHD4lFNdKhttsd4un6psarpNdQLz3YUXmXqBLDgBTKzKORFjyxD/8SH07wk9KhX3GUeUNVMueJcbq4lXpt5t6x4BMHBZOPRQb980IEPlPRJQqrxSOFpVnG83ZS/QOQpiV4BK/kmJ2TfjaNzbHzy4RqboE12tUDf7i1f+2TqbfM9h81LKwVaLsMUjyQ2ERvyPEJxvvrqf6rs5yDY9KU7Ok0Ci/AdHf1rRPDTTz+dVKf2csaOFYV1haEs1Rns0wt9VakOH4itlWlUGstnedEhfPloMqsQLMMfRhN1wjZw2fqS0TSakh6zBsIEYEVnQ6tsJ7FPLzyg6DGrk6qfQgCeZgSZBinVCrnEzbX7zIvcXBU/OgpxPqcRXmUfSLWRGr7aV4D8zRl8ec15Zj3A2oy/EUkPk83gT6bfYRP01vgbbIh29UsR6HtU/vWR16+Pvn7n7PtN3u4NxotnH4+k4LQJGRTnCy+88NJLL5f37Vu6YE8X4fsdt79y27Vr13e/+zz1kcu0bQKlqA+sZwwl8hSlyHoTreBdgu7kIY5l7NOLtx3vhjzrjAhmES02NFxCYBnhW/FHk3SdktjgMOoE91KqeVwCIDPkL2/9IP9osqGFkyh1HiNMEg9nkppgDZtMOezTS1I1leckAFumAZyZEkyMnNas5mYQeXtB+O6ceo672KrQ7VPuXGLnVbauwLssf5k6YEUB85gg34q0ceHq2uTC7/b+sEs6OFC96HtU/r3L11oxFjiJmODZR6aB6cknn3z55b8pXASzr4yQ/bu/+8/s6SJqMwVOfeQyIhgL2MnUtrLKyFOjFPkQnydDniWNtjuOj2DnwNoPeIhjubIoYjviccciylKK3kKO5/zsi8WJ+c/Cj8BqlPANx446wT2cJF6GI3wpa95oEUYTgMwQYGa1kFQfpc5jhEnCWwRTjo6SavYspy0WcPKBtXPHap9yPR3uRgU4MyWYGIXcWdyYTNdgEJnD3aDkHgVzGOHLnVtv7B2Xv4wHTzEePeAm36IU6GD2Sj+duXhvuTWSMYC8sXD1k6m3Px4+0zrVa0JgrUJMmHxjj0YEozjRnfv27fPzE9n69NNPf+97L6B62Vcm72fHtGL3FwvYwRo6mDz2zaVGHQfHD/FkYElDPDHW7O4g+FiZUpxEYFGHpYv6tKItFrCT0qSWS+gt5PjjyyfYwmSVdQkNPwmN5ySVWZmevH+axYn5z8LPpYYnnCRehgN1gmQffOVQ+jgG4VCNyjShoRGUpY4mbxFMGDqiO6YQEwnmTKrAn2iGq9ShJvVpRVsslOpk1AeVBAQ87izeWBjB4M7iluTG8MegTwAAEABJREFUZLp2ZhADOOkZIDCNeRxBgDlcr/A1rnZf/po4wc3jw+TbdWSvdP304vXR168NPdgSvjP78yZL4c3FFcT6jbG5jx4+devYG3fOvv/Zao/fy2zgcPDAYqVhrWqgb7EuoS+f+eLHyFCcqNjR0dEDBw5QmFSZq6hSqrF5TBPUMzr14MGDKNfYJn6FWKMXLGP/7//+v9CX8Y1+n376P/jZLKMV4omxZncHwcfKhPIj8YxGBZqE0qWEhMCiDksX9WlVhjPF2uTln1U2JTTmOXGZRGg8J6nchJXJjwPqBMm+d+G1YBwJ0AxicKTExMuRalSmCQ39evRrRXdMISYSzJlUeEJimgVOkqeExFXqUJP6tPLrzqPVn0+MBs5kyvxZHV/i9AgwZxP3O4s3FkYwuLNomLPr2OaZxqjKysFMBgLTmMdRSQRisaQX9ov8hQKPD0YdZUO+jen+7Y2NC1c/mX4HKYy4RGIiNNlhrVcNo3fvnltil5dd6gdeHZ1bP73Y/L9mS5kAbAj9RWn/gmtKv4VcQnGiYkdH//rll//GiE50p5Uo5yqqlGpsHtMkd9dOBujL+Ea/yGKnNvVV4hmNCjSpFUrXHVU4tCoVlbuHBdYkQDOIwZGSAu0XaIppFjhJvkDLHqZwIHAmU4bZ5dFdN5oQe8Cq4mkW9Nu0DBOpsYPbR/KXMWBm8AqCviHf9oTERGiywxqo4bWJeSOIkaRlRPf52gaWEbv0cmv8jesjrxu9uza5wC4vu9RldFqxzcFXDu1bnGzyHVsxEHUnAiIgAiKQSkAXW0mgv+QvQ8Q7GfoGlUO+Swk1fPf8shHEN44++O7BA206Nscm8doXshjNSmK3GAmbnqhDTZOQuVi4/oXS/d3eH2IZsUsvGxeudkPvhufArlNjfBLKG3y4UHkREAEREAEREIGOEeg7+cv4oW9QOWgd8t1OaGKSkcXrpxdJ7BYjYdMTdahp0saFq1jontK1xv3hPYN75197bOaoVZ7hVFVFQAREQAREQARaQqAf5a8ZGrROq78KbKLQMT+B7YcP/MXyicHxQ/lNyYIIiIAI9CMBxSwCbSPQv/KXkRoYG35ipdJ/mpJOlRpFYOfJZx9fPqEv+zZqUOSMCIiACIiACJRKoK/lL2QfGRrcvzjZD1+EINiyU7vsmy887J59qV1uy1sREAEREAEREIGcBPpd/hp8j80cfXzpBB+Cm1MdO09g4MiDjX994aHzA60ARaAyAupIBESgRQQkf78cLPOLEHwU/uW5/q+jBNj03X3mRbb82fjvaIgKSwREQAREQAREII2A5O/XdNBDfBS+/+LktoNDX5dmy6l2owmw6fsXyyd2Tj3XaC/lnAiIgAiIgAiIQJkEJH9tugNjw4+jkE4+a1/QeZsJBJu++iu3Ng+jfG84AbknAiIgAu0gIPkbM05mG1jfBo5B086iwVcOadO3nUMnr0VABERABESgeAIlyN/inazH4vaRA2wD7z7zIhuH9XigXnMT2HZwaP/Fyb0Lr2nTNzdLGRABERABERCBjhCQ/O0xkDunnmPjcMfxkR71dLlhBHhp2XVqjBeYgbHhhrkmdzpNQMGJgAiIgAg0noDkb+8hYuNw6NwxNhEHjkhI9cbVhBq8rvDS8tjM0UeGBpvgj3wQAREQAREQge4TaE+Ekr+uY8Um4v7FyaG5cT5Pd22jepUT4BWFFxVeV3hpqbxzdSgCIiACIiACItACApK/2QZpx8ToEyvTfKrOZ+vZWqp2yQR4Ldk7/xqvKLyolNyVzPcgoMsiIAIiIAIi0GQCkr8+o8On6hLBPuDKaYPwZVeeEdG/4lYOYFkVAREQARFwJaB6rSAg+es5TI8MDUoEe7IrrlkgfNmVL86qLImACIiACIiACHSZgORvrtGVCM6FL0fjxgvfHLGpqQiIgAiIgAiIQJkEJH8LoBuIYD6CR5YVYFEmkgkMHBneO//aEyvT2vFNhqQrIiACIlAjAXUtAk0nIPlb2AghghFkyDLEGRKtMLsy9BWBHcdHHl86sX9xUt/x/QqJ/l8EREAEREAERCAzAcnfzMh6NkCcIdGe+HAaudbXPxDRk5RbBTbUd50a+8tbPxg6d2z7yAG3RqolAiIgAiIgAiIgAvEEJH/jueQv3TY8hFxjM3hobnz7YYk2H6KDrxxiKx2Gj+kfsPDhpzYiIAL1EVDPIiACDSYg+Vvu4JhvRDy+fILN4J0nn2Ujs9z+OmGdt4XdZ15ku3fvwmtspXciJgUhAiIgAiIgAiLQFAKSv+WORGCdzeDdsy+xkcl2pr4UEWAJZ3g34A3h8aUTvC3snHqON4fwVeVFQAREQAREQAREoBACkr+FYMxghO3MoXPHDqz9wOhgNF+Gxl2syl6vUb28G/CGoG/3dnGQFVOfElDYIiACItBMApK/tY2L0cFoPvY70X+owNpcqaPjgSPDu8+8+MSH0+z1SvXWMQLqUwREQAREQAT6lEAF8rdPybqHzX4n+g8ViBYcmhvfcXykq1vCSHyEPtvef3nrB/sXJ3dOPbdteMgdlGqKgAiIgAiIgAiIQH4Ckr/5GRZmAS24Y2J06NwxtoQ7I4XDkvfx5RMIfba99b3ewiaNDDWdgPwTAREQARFoHAHJ38YNiXEoLIUf7JVenNx1amzwlUPN3xhG77KBvfvMi/svTj55/7QkrxlQHUVABERABESgzwg0N1zJ3+aOTeAZe6UDY8OPzRzdu/AaG8NGDaMvd558duDIcL2CGLGLD0jzobnxQO+ygb1z6jl8DkJQRgREQAREQAREQAQaQkDytyEDkcENo4bRl7tnX9q/OIkgfrDJunQC9YkmRomy+YokJWUwmlr14T2DWCMhuLFvlO4TH04/6Hf5wb9CjDTfMTEqvZtKsY8vKnQREAEREAERaBIByd8mjUYOX7aPHEB9oolRomy+IotJyFOTkKqIYyvtnX8NLWsSeesqp2wzm+YH1h78pRoGEdzYN0p327D+ai3HgKmpCIiACIhAPxBQjI0kIPnbyGEp2imkKuLYSoPjh9CyJpG3rnLKNnPRjsieCIiACIiACIiACNRMQPK35gFQ931CQGGKgAiIgAiIgAg0hIDkb0MGQm6IgAiIgAiIQDcJKCoRaBoByd+mjYj8EQEREAEREAEREAERKJGA5G+JcGV6KwGdiYAIiIAIiIAIiED9BCR/6x8DeSACIiACItB1AopPBESgQQQkfxs0GHJFBERABERABERABESgbAKSv2UT3mpfZyIgAiIgAiIgAiIgArUSkPytFb86FwEREIH+IaBIRUAERKAZBCR/mzEO8kIEREAEREAEREAERKASAhXJ31u3br371X/m/ympJEB18oDAb37zG4M9OIr/Ay76Xx0EmHvBPDQZSupwpMg+dYsVSVO2GkmA+9TcsMGRkkyeUj9oazKUZLJQQWVcMr4FR0oq6LfYLvA58N9kKCm2i7Zbq0j+Qv/7W//73//7f7edXYv8/8lPfrIV//f37t3bIv/lapcIdPJp4H6LdWkoFUtfEch/5+a3UAHwVjjZk0M3ougZZp4KFclftkYsL7/1rW9ZJTotj4D1svHv/t2/K68vWRaBdAKdfBqw2ISj1i0WpqF8NwhE79wXXnghU2hRCw1UAq1wsif2aBRZB6tnF34VmtOqIvkblV/afaxyEkT5V9m7+hKBMIHobOzA08BabBq4qIeHQHkR8CBg3bkek9yywFtiA+/9VjjZc/isKDwGq2cXba9Qkfy11gYmfdvBtct/i7/eAts1fB3z9je/+U04om48DaxbTItNeIiV7wYBa5J73Ln5LVRAshVO9uTQjSh6hpmnQkXy13oRkfzKM2ZZ21ofy9Lc47FFKyURKIRA954GusUKmRgy0nAC1p3r8Y5nWWimEijfySrG2YrCY7Cq8LLWPqqQv9G1oYGfd9Q6CuV2zrz/6db/mvnQKZeCrDeDQCefBtZKA2m9YQJBqUsEondu1kl+69atrQvRT//2b/+2aYha4WRPaPkHq2cXHahQhfy1NuGhhiDjqFQNAV420LvhREk1XauXRAL9eqGTTwOWTGs89YizgOi07QTy37msO+FliHxWAV0Bw1Y42ZND/sHq2UUHKlQhf7U2dGCiKAQRKIRAJ58G1u4v63ohrGREBJpDIHrn5hKvzQmsi55osFxGtQr5a+3Da1/EZWBURwQ6SaCTTwNL/koWdHLq9nlQ1p2rd7wmzwcNlsvoVCF/C18beLNhdE0i7xJn/jp0ZHrkSL6nQepQ06SelV0q8HGGscbRpb6pE3aDhiTsmEsux3Bz8i5NetZhPuDGu+++W5TB2B4J0/TCMbZCpsLA50KsuXdNdyYRjnurlJrYMQY5plRzuYQFk7DpUt/UgaTJmGNYKWLHGORoruY8YsckLOc0ZZpjxxgMR0GhNZPDQZmGsUdaGWscycfW6VmIJzR3SfjZ01pSBasXTpNqhssJynKMknCFovL4Yzrytg8cY4Fjuld0QZ1wovf0JtbVcFuTx6ZVJ3rq7mG0LSU4afrqeaQj6kcTFqzCdFNWZU6t+kkdUTOaQBQ0Jx+tYJVQJ6hvXUo/DVqZTCYnqWxacUzvxeUqwLFjkkv9cB3ahk/JGztJRyq4pDBV8i5NrDo4luSDVQ5Mq23hp6XLXxiRwn577/6C4x//8R+/+c1v7tu3L/g3zMhT8k//9E9WL+Ee8+STOv32t7/9z//8z1HLuIEzXMWxwMmHH36YEsqj9XuWMCf+63/9r1gjzLDBV199NdYByyCzLWhlMkRk1YmeUqdY1AYLDhgUZEgmKLBwNeqDX0lOXFanEIZz2Gfc5pRCLlmVizqFBkyYMHREdyYx+uBiJjA0Hh0VhSWnbzQnhf3/1re+RYmZbMRoguVI7H6QsVYBPTM6ZjiiI0JQ4RitPPVNvAwokZpEnvDxHP+t+rGnVMMIrfDEWOh5ZA7EmnIptHr5yU9+4tKKe8TyCrddGrrUwRS4sM9UCdwDSCaMMGEQTStMmYTBlLnH8JlqwZHeccbFZ+oYn01bc6Rk7969XIpNHh6G7eBYIfMEO6SwZRwz/sceCSpcmTzNrZouswja+M+YMkZBc/IwZ3Zh1kr0QtdcpU5QnwGlhHKrcvSU5kErk3FxEhRZZ1G066CEuJh+xmfjA0dOKeRSUC0lQxSkcAU8xEhS6kkmaRQYF9pafYX7DfLUYRwZFAYiyQ2rHJ+D5iVlSpe/yC/L9fS1wapsTmHH9DKsGQlTGBwpgSxXHSdH0DA9Q6fGLANMF1Zl4sIlBoxqwSVq4gatuBoUmgwllHOVjCnpeWT4sU8irnAvpiG3JQ4wmaK+mQrmGO0ufWuKjjCLn8QStUyJiQKXjP2eRwxiCoM0JCKrvodBy0JwinFYkfCNToNyk3HEZSpzpD4+g4IMp1aikEv0Fe3Iqpn1NGAVHTj6IjS8oo672QKx0C+9M47evsU2xCaWmQlWUB6QsYO1PB5aPg5abDIAABAASURBVKTTYzi4ATlarZIecYwg0wYP8TMaLyV4ztWoQcs+ZKiGEQxal1JO02/8lIbRUUuRa2E7RBQ+Je/tA22DRNTEDgFwMUBBucnQKeVcTcdIQ+5fEtUwaNoGRwgzUgwu1oJCk2Fwo1FQ31xNP2INz8N1IPnjH/84XBLkvT0MLOAVHOgxGmBQJ5qJRked6BygMCVFjUQtQDLFAj6bccR/uFk1scYAMXxUCy5Rk3hpxdWg0GQooZyrZExJ7DF6Nd3J/GMUdsOMF3GRCZebPIVcskI2l6xjNAqrgnUaHaygAnjpFG6wjY4CJYYqN1HQJJrBc2MBa9GrSSUpXiU1yVpeg/zNGhUzDHbpfAkbsoxTz2rUdElMIJ59DHl6ZXwz05Heqc9UIJPShOlCfYx/VSfx/+mamthPrPHFBUzRL8cvzmIO9Bgu5Wmbwp/uikWNY7jXEwvQGDuqhV3NlC8KF50aZ3jVttBxyUoGF/Wtcr9TR1YYBxS4yPRMRWEpyjfsWD7jYTpAR8hYdplp9F4sPZzn4Y7ZIHGLkYLTIGMC6fmAwiCDm1KNS0xOqgWWHTPeX9aM3gjpmiDwx2ro7UBgkIzjQMMHjIw1TaKJWZfn6Rr9uS68ivYSLcElHAuX/7f/9t9iH8g5PaSLYueJY4D0a1I0qKiFaB3TliOVuZ2BQD4lcU8xjiAlUZ/hJpNSnwlJfYwn1YleSnES97CGD0nWTDk28Y2jOY094jZzg/saD2MrBIV0xxpN/aAkmknvK1o/KUbTFxMp2iRcgjM4n1SNcuKiTriJS76Qx0V6R6XL3+hwJrGOdRR2zDB3dgwDYxZryr2Q2UOnUc9jLVCZTqlPJraCVUgs1LcKrVMqcCdbhUmnGEyZXlYUKetW4aiNQcuBpCgo52liKQkKXVKBuIDJUOK5S7/UMfXJ5ExMHvrl6GgHD8GVXrkoLHhVlG/ukyEcWk/IBXoY9JuJXtCKTOwtxnjBkECo4JLoPfY5RqRccrFg1Yn1yqqTdBodNUdrVghxT/6kPuPLDcaoP/G1H3qIeyT6SAFgzqdrVP5Ge4m6hDMWEBZ45G+0Zn4PC58n7sxNONEZErWQNB9wnpslWt9Yto5UBhf1yViXYk+5B6kfe4nCaKdJTmIk5yyiO5NwCf+Z2+a059HUT6kWjSKlMpeig0Uh/uAVfZF3SQCxpjetGBTKyWRNsS5lNdKzftXylxu+p09BBQbAg51Hk6BHMkydTKNOE559DDMZx0RlQkuqjP/Rq8yGf/iHf/jRj37EMbq3hM/RJsa+NSOxY8qtI83p1yrseZrSxBh0v3lMXwRiMu5HfKAvqz5hAiorLrxl6Bkdy1r6KfVZ2NLrpF/FAv3Se3o16yoP3xRcsVgsC9Yp1qKBFOsbXVidOp7iRtQ305ZLTaBnnOHI3OMYTsxPhiNc4pKPbcKgu7SN1klaxaM1oyUQDhdiKvoIClcw+ehY09Bc8jsajFlvE8sNqGLHcoAhy/S4oL4VC16xClhmw6e4YU1gGMZ+7aEQDwufJ/gfDqdnHkRWHctCkhKgWtbbGfLWFLW6tk6pHJ0Dpg69m4w5JjlZyBiZLpg5xItL5tTxSH1rOoUbWlGEL8Xmo4MFH2KMrZxSGG1S+DxM6d3jUunyt7f8SvCaAY7S5KGDpvn1r399//79mzdv8vigxDLA2DN4VqH7aexOKq/79PvTn/6UHnlQplvjuUYdalKfVrG3UJKHzGnrEs2x88tf/hKDbBVwJHDsWz7Q0CrhFIYcwwnfwqcmT7ViUccapC9uM4C8+eabREQmPHYmTAKkmnsi6gJxca/iudU7juEwzJlyjAJuWxU4tXygxD3x+ItOOYYJFHRHpyQcwI2oTcKPFlJCedQlLIAdayZhnFlN5XBiIQmfFu6b9TQI+sI3YkyHTFBB/SBTuIdYpqNYeoGHPH9ipwFtSeFZzSkzqqibC1MWQJ4DOGMGNP2I8zjjl3iihhtaAYYvhfN4Gz4lz+3P0S9hLYoRU9hkLIiOuU0m7BuTikLuI6qZFB1ZU4d7gecqNTkyCaFq6gdHGgZ5k4neO9bQmGrBEeeZq8EpGboLe0sJiY6suefhIawsZ4go5zyxDII6fb7xBCOccLIsMHDhq0E++jDkErTpkdFkgIiFkpRE19ShJvVpBcBoZQtyUMHFyULGKOixjEXHigIImQaL+cN0DTw0GeYqdsws4h4BLyXmUnDkQREGix3LE8bFWEj3h6vc0YHZ8jLlyl9wWK4zNa2SpNPoAHAP8JwKnhqYgiYl0RvJgp7URbScwWPMwuX0wl3EYNAvNxI9MvB0Snm4WpDHGQaYOtSkPq1ozrwJKpgMvVhPQ8op5GYgEySM0Bw7QYnJYN8qxJqlXagZ5Y97lFupWNR4wiPM6oJbhUDgBhDGEefJAIoAIUk4XKXQapV+Wiwu5gyjH+4xcAyHyXMJeriNt+TDCc40D5e45xlxmofr0wugmDNkTDkOwAdW5jQ4xnYaxUJ9fMZCmDDGmdVY5iqJAOmRfskHqVjfrDBNL/RrfMMT8hTiWCxk5lU03mI9pPdYepCBXuAhkxkPLVa0NYmrJmOOvW+uvXsZWawRuGkSHK14rVOq4YbVHYWFJ5iEbYZnUbjcykeHOxqg1STplKHP/0ghCmZLuAuwM6zRcJiQViEOWE/XaCxWhXBHCCZr7LDP2IXrkC/KQ6svLNNXnnkSHUpzq2LZMTla4AkMhLBNOmKMeFIRAtAYMkaHm4XycLUgz7iwrFCHmtSnFc25f4MKJkMvjKnJB0cXJ2lY4CxipAg5cIAMceE8PgdPGyIiCgq5Gk54S/NwiclTbjLBEZtB3iVT1CMr6h6B5JmHLs5nqlOu/GWuWN4wllZJ7ClzwmrLbOAeiA4kJbGTO9Zsz0KeU+E6GGcicheFC8kTBQNJxkqUU59WVjmVuWQVWgFy1bqvCDk66anGrGKCciQfTtF5Hy2JxlI4ahha/RI7D6xo1zhPgDyteFSRz5qKxQVSywHmW6xjFJKsytHRtCrEngIK/uFLsGIKxT4jmOfW1KJ59CFuYcE4kKMOU07CJn1RgVFglobtY7xY32IR5YFcuIcAiaUHGS5ZiZGCnlXIaXieA9CKmpuakMOcaUKiJGrNaks1KzF2Vknhp9HnDK669GI5TyuSS8NonUIeKdbIMhBM+2hfxMujgKN1ickWLqG5FQ53ohWyqU9D/Dd5c6RhbNfFemj6Msec8yQaV1aDUQvcPsa38DEKiodh+IYylWmbdEtSH7ymWnCkMk2CU5OJuhQtibYqdoyYacaZ4MjDIfZZTSEpqGYyUYcpjxZmGqxSH1nxnuB0Talc+cudb8UVnU9WBXPKGJiMOTKhY58X5iq3h2U22q+pmX7k9d1qGHvbGCNWj6aQuYurJm8deVxaJdYpXYefudiJhoyH3//iP4uPZSo4te6E2MlnmYrtNzDYEzVrgPUIo9PYR1Jgkx6DvHumWFyQx2C4d7QIwYZLwvmUS+FqPfMWK+oz6ElAKI/OImuIiYJYsBMkHpqk4NTKMDoIXypg3LpUhm9WF+mQo8Fazcvw0KKHD8Cx+g1Oo9MAnsFVMsXeXBi0EnuiQEhJ1vSwmrucMqOsarGPPqsOp1ZDx1Y0tFIhjxScCY8ss50bzeoo69OVuRG1YJVwisQhBDJBYlmx5gmXyvAQsyblnCf4ZuwEx+jMDy7FZqIWovMB/lY1QEWrGfux5XnWX8xavVNi9UKFAmcRpjBIL0FKfx46MrdsYtyxITVJpT6ycs5D3Cs2lSt/rYcvDx1SzwAYP6sht0F6w/SrPXs0FZiOJmOO2My08vE0jD7UjCmXIzd/uBrWcMCU8PRkUn7zm99k9lhOmgr4iYiBkjkNjhbGqHuFo8bPoHeTYZkJAjElhRyLxWVZw0OQckxKRUVk9cugWw9cy4HoCFoVLINcjc4KCl2SZSq/b9ZsxIcAMvloyhpsfg+tkHGJBYljUooOVrik8JsrOutwmO2olPTtb3+b92UeIEkh9CyPtg3HmNScVtZwZ1qDw2YLeaQAKmyTqRLAxFW68Hi6RiOyeqFH3kysJzatYm9Jq20eD4PQcMAkjKdMEi6lzxNrKF0mgOk3OFoWuLWjTlqgqJDyfABjYNxkIIZZk/c79nQSjGHL9IiTpsRjFlnWsJMSL1eDvsinJCuKTINV7CMr6jAhM9lSUvo8TIna71K58heaYbccRwJG4Vbko3OdwpTk2JFlITpvouMXNLHuVcqzdmrVtwyakLmpeIDyaGYLwYJJj7jH8pz0PXQqWE2MTcqDVDhqyyAxRjsNes+TKRaXNfT4DNtM7mWtj3E6ZXzJBIl+g7xjxmpiYYG/35JQhm/R2ZgOzapvASnDw/z0AB74ad0LlFuDRUl6ClujZni55dQxEVQeBQxnx47C1aKxp491uK2Vt0zBJCtGDAKBY5CMBe6+PE9XhiMwaDLM2DAuTrFvLpkjENgOMHnrWKCHOEZHlv2epziQNE8IJNzcw7hlIfahFEZHdwx0Skd4S51won74tGc+Wr+nk1anOWdROF68xVpKvFSIptj6VhSxdaKmTIl1r1GIVxzdU5hq4fPQ3Q3HmuXKX2uAHX2KVgszjV6lxOqoZ32aRJNlJH3geXRaFtI7jU5Ka15avWMfybtv3z5elchbffH44DGK8GUjwbIT1LTuVcppxTE9pUdBW8tPq751lRuAJmUkqyMQ5cFlWUsfesKxRpMSF7ZUCycPI5af0U6tCklzI+xGbL4C3zwghydbBR72pJf1Fgv7H4vdGj6rPv5w18c2TC/ELI+R9DpJV7mzrEtR8lYFTi3ZR0l0rlLoknA+XM3vkWIZIag8j4vAn6gz4SlBF3QUVCbD4zqJQ4EeFj5PLN963rlEaiUXCy51ArMWWMqtm4WScLImLYhI4QrkezpgVcAHhth7jbas9aRqhYDDsXMpq1nspKR0qjS0ugvXh3D1zytcck/lyl93P8I1w0+QcHlSnlcWJmL4angMwuUp+WinsXMrsBCdi+nTN2WWGJtWCNxXfCRnLoWP9PLmm2/++te/Tv+ghCZRD6MRRaOmYUpKRx21Fu0xxXimSwXi8nDb6h3PGReOmdLWAerdlE7hH64XxUudcAXvfBm+ZXXGCpaHaRhyGR5mpWd5SIBhD6PzigopCWuWA9HnGHqL2x8UKXZiL/EwyUrM2In6gJ/mUtKRZ1e0rzCZpIbR8ijD6JyPtoqWWGDxECDRajgJXpenq2lLfZMJjgEcXgAs56mM/A1qWpliPSx1nliuWoFETy0OVIhO4Gid9IHOOsF6rr9RB6JOWoHnmUXR7tLjBZrVOyXMKI7pKdoqpX7Uq5TKXGK2W/atx0Wp8xAHcqZK5a9FKsn16LSz5q7VkAdNuITmQA+XuOSjt1P6dLQmSnplHLBCsGYJFXom9O5Pv/jPMbpoRNFOYWX1a/lpXU1HHe2xJxbLfoGn7rg83M46+o5xZYJNLKjRAAAQAElEQVSPzZ6Pv2hotAqnnhWCyoX7xqMzMB7NQJgULu857Qv3MB0O3VkSCg/DN1Q4bwKhicnEHtNvrqAJvaDPUGk/Sv0veuulhxPYtzLRKPAzJRC2mS0sGIw+eSh0SVGfo3G52OlZx/1xETbFWIRPyUMGn0lQ4jRIYPTbCQssZPUQ30qaJ8QYeOWSgYZVLToffvOb31h10gfaejikV8ay5bOLA9E62ElP7mMUZdIzhKwhG1etwE1h0pFZal1Kbx6d5Mw6ywIlJc1DqyOP03Llr0UTlKSeXkannUU5bIE3MMsmb9hWv+H6SfmoNI+6EW5rTd/0uWt5iJ1o/aTuiIWbignE0zMsd7CZLiCoQEdBirUfLcyDOivDwDePTNRzY8QDl4fb1ugnOWNcSjpGW0V1Q9CWS9GhIdiggslY8wo/UyYJBr/5zW+iV1wI4IDpInrkEqas8qhvVgXmp/VADypw6dVXXw1OTYb72mTMsV56EOPJYzwJjtynQZ5M1MMoJaqZhDWiNnlzJN4khpSzqFAhJaGNjZ2cRzqyLBD797///ejY4T/lsTFa09IymHJKX9bVKFWrQuxpUitIMmoeT1fTC82jfLjjGE3Lc0YqHUIZHhr36DolucwT7Jh4zZGh58Fi8uEjITP6hB8uJE85x3CKButSJ2zBciCdLTMz3JZ8tL6LA1G3MUWCT9ZZ5NIdlsPJCjnFmXCrTIMVtcmAhq2F80xyCyzTDBThOiZPIbcJV1OSyzw01go8lit/ozRBZg08g8oCHA4pOjW5o6LDgB3WSJbecFtAMxHDJY55Zkm4ZtTz8FXyuM0xSGFhGhQGGasy5dEYY3tkuphHc7g+gUPj29/+NuGTwVpssqZmrP2wWWMkD+qsDE2PXx+z5GLD8cOV1W34WwMa60zPaKLwMcsNEm3IKMeWR/uNzkMaWjPB2Memue/IIII5mnJzrNE3hgMVBWTjiTkyspZL1inVKqOHh9x9FlWe7xb8qId5bi4CzJSgYdXn2WiVuJwShRUXrRgdxggITBsSE4k8CTJcjaboRI3WiS2xDHrbiW3IpPJ+ugbeRi1Dw3IbgPQVNInNRO1QjVb5PcROSnKZJ1HfWHqshizETABiJ1ndWTSi1qjvUodqQbJ6h3BwKZqxKlOBWc0xnFwciPXcb4xcugu7xx1nRRHrDE2i5e6DFcVS2SPLio5A/J5XNHRP5crf6KRkzWCt5Q7hoUnihuGUDOWB0ywk0chpwgOXmswbhoRFnYZkglYmwxZptK25lH4MO0DN6ByiMEj4EORNJjpvTLk5Roc2SoaoTeXwkWcKie5MIm8Ch4apRobJzb1hToMjJaTglEysh3QaxYVNP9SZGOJSnoTn0ebwIRlWHMm74MrqtlUfN2LZUp6eaBWdBvjMxA7mOWPBKcdYU9FZGjXINAhGEyYkjFs2qcOtxDHopQzfokPGfYFvREfvxjFOSWFPcIkwWWPIhFMZHsbS4xnFLLI8xHPLGZ484RLyxFvgzcWouSeQUhkfggQuMAanmTJJGzPcCHREoi/yKTbpPeVqyiXLrHcIjEW0F+41EhPPJPIMdPjWILTYp2vYVKzlcAXmQHRuhCuY/AM7Jhc64hLJuMeRfE8PGQv3RIBUDnX4ECMVhRy9LxgXQAGH5hzJ45i5KTjiZ9gmlcOndBE+NXmrTtQHU80cQWEywTHWZnAVl4K8ycRGZC6ZY6zBosaILjLFG61PSayHlMeGxgAxTD0HiwCZrhgJJyYJz2Tagp3HIAONNTLhOuSZ5OG21HdPdEFljASJ6NLnQFAzT6Zc+ctGbJiIcZS1jVAJmBRMAsiaq+YYXe0opw5NGAkGkhsMOxSGEwPA+IVLHPOYIoUrQz98auUDt4Py9Pp4HtQkE1uZWUviajjhlQmZqEnMvGjgTETLPhaiHiZNpqJQ4yqJroMUG2ZwNWcGViTLCA5kxUUTUthOT7ejbHs2CdsP52Ph87A2UTDPuVM4DTcJ8rEDyvzviQXjlk1uUu4djoFxMoX7xtMgCgr4xEikTG8ci85kvPrpT3/KEZesVLiHsfTolJsu3cM333zT3UNiJFLixSaWIUAX4cRY4Em4hDxN3BNILbOxrDDrkphpSQrYpTl1sMAxayIEUrhVdP6Er6bkuSlIVgWMg5SBMMn96Rq2g0vp0UGeOuEmsXncI1mXPDwkIvfkOE+4c2OnN0sPfXG0nidhs4RACscVxUUFUrhOOrGsT2DuuHTj9E4K14k6yVUGiEQmnGgIBDOFOLrMIpqQwkbS46Wme8h5BouOmK4crQRAE6P7I4v67ik8YUzXsW6YSwUey5W/3DOOYVijS6vY+ZcSOWsGA59SIeWS1Ts103u35m7P+tbTIWmu//jHP4YY1jIlAo8ultGIkjotCnW0x3SGmWKMrUzg+XF5uJ119GOdN4U8TOFv8lmPSQOKUsmEhcroy+hgleEbQ5YpTLz65S9/iYexrcrwsCh6xmEGlxBM3vEIouhzjOXHsXlsNUBFbcbWTCokEFLS1XA58Ub7Spqr4YbRvMe9GTUSlAA2aSIFdaIZWkWfrla1lAqQd+SGTfrK6WFJ8wSv3KMgENa7wBOXQXSpg9kgZX0C40/Qlkx0Nro7kH+McMC9Oyqb5B5ynsGiLwaaW5iMe4KJdcsHo+9uJFyTu8YyGL5aYL5c+Yuj0HSJJDohYpdkDEYTs5nKLr1E25qSaO/YNJdij1b9ntPFuv0Y3VizdEogTN/Yq9FC7CAOYgO3eqQtxjnGJjrtGYJpiBEqx/ZoMaE+lTmWl7CPMzlxebhtNWEU8sSI3orladkkTOINFyYNGeXuWHD+17/+NU3CloN84b7RHc/KwH56BlVBIFbUVpPCPQQFnULb6ij2lHBS6JkmWMOmyacfiZTKsZMhejunmwpfpXc2p8MlfnlQYyedDE97QrC8hZJ7j+Ga1o3GJRBx9Eu0xbd0/8OWcTvp6RquRp6JyjGa6Mt9ttM8v4cWeWy6p/R5wsjGzsyofRNFwMRlEF3qhDuy6uN5+Go0b2FhZK06lkGuEgXHaKI8/yxy7y5wwGoSDSGoScZ7sGhLIsCeSKlGMjSiE8MCTk33RNc8Z9zr56lZuvzFOR4BjAeZpMRjIkqQQoahZ0Meyjyk0mdDUr9BuTVadM24BlejGWsuMmDROkFJ9E0oxTimWFCjNAJrJkO8TBH4UN+UWEfLQ+pbFcKnxIupnKizMgw74J0n/Jy4PNy2BjRlNB3j4gZhGjMKSfUJkwGyrlJolQSnXKI+x6AkmsFtM4VS+qVV4b4xt7lh6R3jSYmr+I976b6Z5oV7CDd652jsxx7xkH6p1tNDKlCt2Jsr1qXYQvqld3yIvZq1EE3D7UbgZMI2ocGwcslM46JuEI97Mz0ixhQncTW9Gk9L5h7cqJ9e01w14Zt8cIQPFrgUlLhk6DGPhxYxlx5NHZd5wrib8TVNokeCpQ7+AzC4GnWJasFVk7HqgC5ax9Q0R2t1A5opjz1as5E6UeOWA7F1KDSJ7ogxzyyyuusZL/1aUURDoE44MRAeg2Us4A9TlylhTmOP1ME+T/LwWAc1rQCD8p4ZOqVrjPesWUiFKuQvjkLKzJjwsBEkj1HGiUuETbUv0tcHKpiGXA1Txki4IdW+buOVs0aL+Z1ihsrWJxH4k1LfulepmV6fcAIm0agNEKYIBDAVm3DP6jRsJ7YJnRrL3qjBEraczjBcM2cez/Pgyuq2VR/n00eTCi4J7NwFDAEjRUSmCZYZZaLjKQNPq2uummqxR+rTilWcxzT5oA726Yv5Q3cYD8pTMtSncrG+YZC4cCAcRdg38ikuWZfK8DCWHiThCVX8J2O5kXTKgEKPJvgZjovYIQAHc4lqSRas2zmpWlCOZdzDLP2mmA3qu2ewhmUI3Lx58/5X/9ERUdApdn7yk59wDCeghU/d89aE97YT7hH/cRWHY8cCXFzi7mBcwq165rFJW0OAI4iYP34O5/Gw7HkCNOIi0vA0Jk85UwJ0BG6xsgaRylYFTq066dyozAJHqyABPMhHM1Em0frYDDeMdTJcIc8YYcfqLj3eaH1KoiFQaCUGJetgBRYIkFFmQDESpkG/3BrMdnOJakGTcCbKPHw1mscsMwebdJpkM9oqf0lF8hdHidBQ++qZeZ8HKPcMYacHTEOg8EgKGoLJpSGdOqawcXrhNKUh/lAnnJgiKfW5Gq5Mvud0xxq9xEaNNS5RISVFV6DwDE5piOXYTnuOETaBRmhB4pTCylKS5z1x4WfgMxlO032mI6qFE12kN3G8yl2AKRzgvjD2g3luLJjC4IgnpjzlaB5VPASDVthniB3nQ2C5DN+YVNzFxJjTN+NkGR5G6UGShxjlptNMR8YL8vAP4iV2CMAB50OmYrLhVkHzlAyW8ZMeY2x5FVkLdoqNf/7nfw5fJTQ/XBixouaUwkISZGLHghuQS35d0BbsDArH/PBxw8NDEOGAe/JwFceINNwReVxNGmWuhv3hNIqXwp51glY4EK5MHn+Cq9EMV6kTTtH1N5MDQRd4QuDhtvDkjqZHLgXVoplwExzjNFonXII1qoUTXYQrJOVpSE3sB23J43PSYFl2aE5lmgTNTYA9H1nhJkHblAxm898ylvMup9XJXxdvVCc/Ad6M/+mf/ilsh0kcveHDFZQXAREQgVgCfOr6/e9//9vf/jYPltgK4UKePNQPl7DQooDDJcqLgAiIQDKB6q5I/lbHupqe/vEf/9HaquFdrZqu1YsIiEBnCPAYefXVV9G+KFq0L/n00NC+PHzCdRC+bD6FS5QXAREQgYYQkPxtyEBkdoMPGVmfws1YosyPDoYL2fqV/A0DUb4FBORi3QSQvNaP21OCFLaeOcZNChHHlvblEtqX5w8ZJREQARFoGgHJ36aNiJM/aF+ULusTRzZdfvKTn3DkA0rKrfY//vGP2YOxCnUqAiIgAikEvvWtb0WVKwqYZw5Kl6cNjxpOyXBKIY8gyxpv3chfq1CnIiACvQmoRiUEJH8rwVxoJ+zysuoYkyxCbLqwAnFkD8YUBkeWn6x/5BS0VUYERKBvCfDO/KMf/Sg2fJQuTxtevNkMJsNptBralxfvaLlKREAERKAhBCR/GzIQGdxA8kaVbrQ9K1DSAhatrJKmEZA/IlAvgb/927/lGeLhA62kfT24qYkIiECVBCR/q6RdQF8IX3ZcehpC+GoF6klJFURABFII8AzhE6SUCtYl9ozffPNNWlnlOhWBrARUXwTKJiD5Wzbhgu0jf1ljUox+61vf+ulPf5pp0UqxpksiIAL9TIAXaRRt9HvAFhMeStT89a9/zZ6xdUmnIiACItBAApK/DRyUNJdeeOEF1hhWGjLheiw/LDwsVL/85S+tS+FqrcrLWREQgfoJ8GDhmcOzhZdq69mCLOYq271U4CpPofrdlQciIAIi4EBA8tcBUsOqsMaw0rDFG/5nVG7evMn6xFLUMGfljgiIQBcI8Gzh1fjs7QAABkNJREFUrdt67KB6eez8wz/8Aw+lLgTZrBjkjQiIQIkEJH9LhCvTIiACIiACIiACIiACTSMg+du0Ednqj85EQAREQAREQAREQAQKJSD5WyhOGRMBERABESiKgOyIgAiIQDkEJH/L4SqrIiACIiACIiACIiACjSTQAvnbSG5ySgREQAREQAREQAREoJUEJH9bOWxyWgREoE8IKEwREAEREIHCCUj+Fo5UBkVABERABERABERABPISKK+95G95bGVZBERABERABERABESgcQQkfxs3JHJIBERgKwGdiYAIiIAIiECRBCR/i6QpWyIgAiIgAiIgAiJQHAFZKoWA5G8pWGVUBERABERABERABESgmQQkf5s5LvJKBLYS0JkIiIAIiIAIiEBBBCR/CwIpMyIgAiIgAiIgAmUQkE0RKJqA5G/RRGVPBERABERABERABESgwQQkfxs8OHJtKwGdiYAIiIAIiIAIiEB+ApK/+RnKggiIgAiIgAiUS0DWRUAECiQg+VsgTJkSAREQAREQAREQARFoOgHJ36aP0Fb/dCYCIiACIiACIiACIpCLgORvLnxqLAIiIAIiUBUB9SMCIiACxRCQ/C2Go6yIgAiIgAiIgAiIgAi0gkAL5W8ruMpJERABERABERABERCBRhKQ/G3ksMgpERABEYgloEIREAEREIHcBCR/cyOUAREQAREQAREQAREQgbIJFGdf8rc4lrIkAiIgAiIgAiIgAiLQeAKSv40fIjkoAiKwlYDOREAEREAERCAPAcnfPPTUVgREQAREQAREQASqI6CeCiEg+VsIRhkRAREQAREQAREQARFoBwHJ33aMk7wUga0EdCYCIiACIiACIuBJQPLXE5yaiYAIiIAIiIAI1EFAfYpAXgKSv3kJqr0IiIAIiIAIiIAIiECLCEj+tmiw5OpWAjoTAREQAREQAREQgewEJH+zM1MLERABERABEaiXgHoXARHIQUDyNwc8NRUBERABERABERABEWgbAcnfto3YVn91JgIiIAIiIAIiIAIikImA5G8mXKosAiIgAiLQFALyQwREQAT8CEj++nFTKxEQAREQAREQAREQgVYS6ID8bSV3OS0CIiACIiACIiACIlALAcnfWrCrUxEQAREohICMiIAIiIAIZCYg+ZsZmRqIgAiIgAiIgAiIgAjUTcC/f8lff3ZqKQIiIAIiIAIiIAIi0DoCkr+tGzI5LAIisJWAzkRABERABEQgCwHJ3yy0VFcEREAEREAEREAEmkNAnngRkPz1wqZGIiACIiACIiACIiAC7SQg+dvOcZPXIrCVgM5EQAREQAREQAQcCUj+OoJSNREQAREQAREQgSYSkE8ikJWA5G9WYqovAiIgAiIgAiIgAiLQYgKSvy0ePLm+lYDOREAEREAEREAERKA3Acnf3oxUQwREQAREQASaTUDeiYAIZCAg+ZsBlqqKgAiIgAiIgAiIgAi0nYDkb9tHcKv/OhMBERABERABERABEUglIPmbikcXRUAEREAE2kJAfoqACIiAGwHJXzdOqiUCIiACIiACIiACItAJAh2Uv50YFwUhAiIgAiIgAiIgAiJQCgHJ31KwyqgIiIAI1EJAnYqACIiACPQkIPnbE5EqiIAIiIAIiIAIiIAINJ2Au3+Sv+6sVFMEREAEREAEREAERKD1BCR/Wz+ECkAERGArAZ2JgAiIgAiIQBoByd80OromAiIgAiIgAiIgAu0hIE+dCEj+OmFSJREQAREQAREQAREQgW4QkPztxjgqChHYSkBnIiACIiACIiACCQQkfxPAqFgEREAEREAERKCNBOSzCPQiIPnbi5Cui4AIiIAIiIAIiIAIdIiA5G+HBlOhbCWgMxEQAREQAREQARGIEpD8jTJRiQiIgAiIgAi0m4C8FwERSCEg+ZsCR5dEQAREQAREQAREQAS6RkDyt2sjujUenYmACIiACIiACIiACGwhIPm7BYdOREAEREAEukJAcYiACIhAPAHJ33guKhUBERABERABERABEegkgT6Qv50cNwUlAiIgAiIgAiIgAiLgRUDy1wubGomACIhAKwjISREQAREQgQgByd8IEhWIgAiIgAiIgAiIgAi0nUCy/5K/yWx0RQREQAREQAREQAREoHME/n8AAAD//1vPPKQAAAAGSURBVAMA9TjA2oBe1VcAAAAASUVORK5CYII='

HTML = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Comprobantes Pendientes SAP — {periodo}</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

:root {{
  /* Paleta oficina: blanco roto, grafito, azul petróleo, rojo señal */
  --bg:        #F8FAFC;
  --surface:   #FFFFFF;
  --surface2:  #F1F5F9;
  --border:    #E2E8F0;
  --border2:   #CBD5E1;

  --ink:       #0F172A;
  --ink2:      #334155;
  --ink3:      #64748B;

  --navy:      #1E40AF;
  --navy2:     #1D4ED8;
  --navy3:     #2563EB;

  --red:       #DC2626;
  --red-bg:    #FEF2F2;
  --red-bdr:   #FECACA;

  --green:     #16A34A;
  --green-bg:  #F0FDF4;
  --green-bdr: #BBF7D0;

  --amber:     #D97706;
  --amber-bg:  #FFFBEB;

  --mono: 'JetBrains Mono', monospace;
  --sans: 'Inter', sans-serif;
  --radius: 6px;
  --shadow-sm: 0 1px 3px rgba(0,0,0,.06), 0 1px 2px rgba(0,0,0,.04);
  --shadow:    0 2px 6px rgba(0,0,0,.06), 0 1px 3px rgba(0,0,0,.04);
}}

* {{ margin:0; padding:0; box-sizing:border-box }}
body {{ font-family:var(--sans); background:var(--bg); color:var(--ink); font-size:13px; line-height:1.5 }}

/* ═══ HEADER ════════════════════════════════════════════════════════ */
.hdr {{
  background: var(--navy);
  border-bottom: 3px solid var(--red);
  padding: 0 24px;
  display: flex;
  align-items: stretch;
  min-height: 90px;
}}
.hdr-brand {{
  display: flex;
  flex-direction: column;
  justify-content: center;
  padding: 20px 32px 20px 0;
  border-right: 1px solid rgba(255,255,255,.15);
  margin-right: 32px;
  min-width: 220px;
}}
.hdr-eyebrow {{
  font-size: 9px;
  font-weight: 700;
  letter-spacing: 2px;
  text-transform: uppercase;
  color: rgba(255,255,255,.65);
  margin-bottom: 6px;
}}
.hdr-title {{
  font-size: 18px;
  font-weight: 700;
  color: #FFFFFF;
  letter-spacing: -.3px;
  line-height: 1.2;
}}


/* KPIs en header */
.hdr-kpis {{
  display: flex;
  align-items: stretch;
  flex: 1;
}}
.hdr-kpi {{
  display: flex;
  flex-direction: column;
  justify-content: center;
  padding: 0 28px;
  border-right: 1px solid rgba(255,255,255,.15);
}}
.hdr-kpi:last-child {{ border-right: none }}
.hdr-kpi-lbl {{
  font-size: 10px;
  font-weight: 700;
  letter-spacing: 1px;
  text-transform: uppercase;
  color: #A8C4E0;
  margin-bottom: 5px;
}}
.hdr-kpi-val {{
  font-family: var(--mono);
  font-size: 26px;
  font-weight: 600;
  color: #FFFFFF;
  line-height: 1;
}}
.hdr-kpi-val.danger  {{ color: #FFA09A }}
.hdr-kpi-val.success {{ color: #7EEAAA }}
.hdr-kpi-sub {{
  font-size: 11px;
  color: rgba(255,255,255,.6);
  margin-top: 5px;
  font-weight: 400;
}}

.hdr-periodo {{
  display: flex;
  flex-direction: column;
  justify-content: center;
  align-items: flex-end;
  padding: 20px 0 20px 28px;
  border-left: 1px solid rgba(255,255,255,.15);
  margin-left: auto;
}}
.hdr-mes {{
  font-family: var(--mono);
  font-size: 17px;
  font-weight: 600;
  color: #FFFFFF;
  letter-spacing: 1px;
}}
.hdr-fechas {{
  font-size: 11px;
  color: rgba(255,255,255,.65);
  margin-top: 5px;
  text-align: right;
}}

/* ═══ BARRA PROGRESO ════════════════════════════════════════════════ */
.prog-wrap {{ padding: 16px 36px 0 }}
.prog-card {{
  background: var(--surface);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 12px 20px;
  box-shadow: var(--shadow-sm);
  display: flex;
  align-items: center;
  gap: 20px;
}}
.prog-lbl {{
  font-size: 10px;
  font-weight: 600;
  text-transform: uppercase;
  letter-spacing: 1px;
  color: var(--ink3);
  white-space: nowrap;
  min-width: 110px;
}}
.prog-track {{
  flex: 1;
  height: 8px;
  background: var(--border);
  border-radius: 4px;
  overflow: hidden;
  display: flex;
}}
.prog-reg  {{ height:100%; background: #16A34A; width:{pct_reg}% }}
.prog-pend {{ height:100%; background: #DC2626; width:{pct_pend}% }}
.prog-legend {{
  display: flex;
  gap: 20px;
  font-size: 11px;
  white-space: nowrap;
}}
.prog-legend-item {{
  display: flex;
  align-items: center;
  gap: 6px;
  color: var(--ink2);
}}
.dot {{
  width: 8px; height: 8px;
  border-radius: 50%;
  display: inline-block;
  flex-shrink: 0;
}}

/* ═══ TARJETAS POR TIPO ═════════════════════════════════════════════ */
.tipo-cards {{
  display: flex;
  gap: 10px;
  padding: 10px 24px 0;
  flex-wrap: wrap;
}}
.tipo-card {{
  background: var(--surface);
  border: 1px solid var(--border);
  border-left: 4px solid #ccc;
  border-radius: var(--radius);
  padding: 11px 14px;
  box-shadow: var(--shadow-sm);
  min-width: 100px;
  flex: 1;
}}
.tipo-card-lbl {{
  font-size: 9px;
  font-weight: 600;
  text-transform: uppercase;
  letter-spacing: 1px;
  color: var(--ink3);
  margin-bottom: 6px;
}}
.tipo-card-num {{
  font-family: var(--mono);
  font-size: 22px;
  font-weight: 500;
  line-height: 1;
  color: var(--ink);
}}
.tipo-card-total {{
  font-family: var(--mono);
  font-size: 11px;
  color: var(--ink3);
  margin-top: 4px;
}}



/* ═══ SECCIONES ══════════════════════════════════════════════════════ */
.sec {{ padding: 18px 36px 0 }}
.sec-hdr {{
  display: flex;
  align-items: center;
  gap: 10px;
  margin-bottom: 10px;
  padding-bottom: 8px;
  border-bottom: 1px solid var(--border);
}}
.sec-title {{
  font-size: 11px;
  font-weight: 600;
  text-transform: uppercase;
  letter-spacing: 1px;
  color: var(--navy);
}}
.sec-line {{ flex: 1 }}
.badge {{
  background: var(--ink2);
  color: #fff;
  border-radius: 3px;
  padding: 3px 10px;
  font-size: 11px;
  font-weight: 600;
  font-family: var(--mono);
  letter-spacing: .3px;
}}
.badge.red {{ background: var(--red) }}

/* ═══ FILTROS ════════════════════════════════════════════════════════ */
.fbar {{
  background: var(--surface);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 14px 16px;
  margin-bottom: 8px;
  box-shadow: var(--shadow-sm);
}}
.fbar-row {{
  display: flex;
  flex-wrap: wrap;
  gap: 10px;
  align-items: flex-end;
}}
.fbar-row + .fbar-row {{
  margin-top: 10px;
  padding-top: 10px;
  border-top: 1px solid var(--border);
}}
.fg {{ display: flex; flex-direction: column; gap: 4px }}
.fg label {{
  font-size: 11px;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 1px;
  color: var(--ink3);
}}
.fg input, .fg select {{
  border: 1.5px solid var(--border2);
  border-radius: var(--radius);
  padding: 8px 12px;
  font-size: 13px;
  font-family: var(--sans);
  color: var(--ink);
  background: var(--surface2);
  outline: none;
  transition: border-color .15s, box-shadow .15s;
  min-width: 140px;
}}
.fg input:focus, .fg select:focus {{
  border-color: var(--navy3);
  box-shadow: 0 0 0 3px rgba(37,99,235,.12);
  background: #fff;
}}
.fg-wide input {{ min-width: 210px }}
.fg-num input  {{ min-width: 100px }}
.razon-wrap {{ display: flex; gap: 6px; align-items: flex-end }}
.razon-wrap input {{ flex: 1; min-width: 200px }}

.btn-buscar {{
  padding: 7px 16px;
  border-radius: var(--radius);
  font-size: 12px;
  font-family: var(--sans);
  font-weight: 700;
  cursor: pointer;
  border: 1.5px solid var(--navy3);
  background: var(--navy3);
  color: #fff;
  display: flex;
  align-items: center;
  gap: 5px;
  transition: all .15s;
  letter-spacing: .3px;
}}
.btn-buscar:hover {{ background: var(--navy2); border-color: var(--navy2) }}

.fbar-actions {{ display: flex; gap: 8px; align-items: flex-end; margin-left: auto }}
.btn {{
  padding: 7px 16px;
  border-radius: var(--radius);
  font-size: 13px;
  font-family: var(--sans);
  font-weight: 600;
  cursor: pointer;
  border: none;
  transition: all .15s;
  display: flex;
  align-items: center;
  gap: 5px;
}}
.btn-clear {{
  background: var(--surface2);
  color: var(--ink3);
  border: 1.5px solid var(--border2);
}}
.btn-clear:hover {{ background: var(--border); color: var(--ink) }}
.btn-excel {{
  background: #1A5C35;
  color: #fff;
  font-weight: 700;
  box-shadow: 0 2px 6px rgba(0,0,0,.3);
  border: 1.5px solid #236B40;
}}
.btn-excel:hover {{ background: #145029 }}

/* CHIPS */
.chips {{ display:flex; flex-wrap:wrap; gap:5px; margin-top:8px; min-height:20px }}
.chip {{
  background: #EFF6FF;
  border: 1px solid #BFDBFE;
  border-radius: 3px;
  padding: 3px 9px;
  font-size: 10px;
  font-weight: 600;
  color: #1E40AF;
  display: flex;
  align-items: center;
  gap: 4px;
}}
.chip-x {{
  cursor: pointer;
  color: #60A5FA;
  font-size: 11px;
  line-height: 1;
  transition: color .1s;
}}
.chip-x:hover {{ color: var(--red) }}

/* RESULTADO */
.res-bar {{
  display: flex;
  justify-content: space-between;
  align-items: center;
  margin-bottom: 6px;
  padding: 0 2px;
}}
.res-bar {{ background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 10px 16px; margin-bottom: 8px; box-shadow: var(--shadow-sm); }}
.res-count {{ font-size: 12px; color: var(--ink3) }}
.res-count strong {{ color: var(--ink); font-size: 15px; font-weight: 700 }}
.res-totals {{ display: flex; gap: 28px; font-family: var(--mono); font-size: 13px }}
.res-item {{ display: flex; flex-direction: column; align-items: flex-end }}
.res-lbl {{ font-size: 9px; text-transform: uppercase; letter-spacing: 1px; color: var(--ink3); margin-bottom: 2px; font-weight: 600 }}
.res-val {{ color: var(--ink2); font-weight: 600; font-size: 15px }}
.res-val.igv-v {{ color: var(--navy2) }}
.res-val.tot-v {{ color: var(--red); font-weight: 700; font-size: 17px }}

/* ═══ TABLA ══════════════════════════════════════════════════════════ */
.tbl-wrap {{
  overflow-x: auto;
  border-radius: var(--radius);
  box-shadow: var(--shadow-sm);
  margin-bottom: 24px;
  background: var(--surface);
  border: 1px solid var(--border);
}}
table {{ border-collapse: collapse; width: 100%; min-width: 860px }}
thead th {{
  background: var(--ink);
  color: rgba(255,255,255,.9);
  font-size: 10px;
  font-weight: 600;
  text-transform: uppercase;
  letter-spacing: 1px;
  padding: 11px 13px;
  text-align: left;
  white-space: nowrap;
  border-right: 1px solid rgba(255,255,255,.1);
}}
thead th:last-child {{ border-right: none }}
thead th.r {{ text-align: right }}
thead th.igv-h  {{ background: #1D4ED8 }}
thead th.tot-h  {{ background: #991B1B }}

tbody tr {{ border-bottom: 1px solid var(--border); transition: background .1s }}
tbody tr:nth-child(even) {{ background: var(--surface2) }}
tbody tr:hover {{ background: #EEF2FA }}
tbody td {{ padding: 9px 13px; vertical-align: middle }}

.mono  {{ font-family: var(--mono); font-size: 12px }}
.muted {{ color: var(--ink2) }}
.num   {{ text-align: right; font-family: var(--mono); font-size: 13px }}
.igv   {{ background: #F0F5FF !important }}
.money {{ font-weight: 600; color: var(--red) }}
.total-cp {{ background: var(--red-bg) !important; font-weight: 600; color: var(--red) }}
.proveedor-cell {{
  max-width: 240px;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
  font-size: 13px;
}}
.col-fecha  {{ width: 88px  }}
.col-tipo   {{ width: 120px }}
.col-serie  {{ width: 52px  }}
.col-ncp    {{ width: 68px  }}
.col-ruc    {{ width: 100px }}
.col-importe{{ width: 14%   }}
.col-total  {{ width: 14%   }}
.col-mon    {{ width: 44px  }}
.empty-td   {{ text-align:center; padding:36px!important; color:var(--ink3); font-size:12px }}

/* Tipo badges */
.tipo-badge {{
  display: inline-block;
  padding: 1px 7px;
  border-radius: 3px;
  font-size: 9.5px;
  font-weight: 600;
  letter-spacing: .3px;
  white-space: nowrap;
  border: 1px solid transparent;
}}

/* ═══ FOOTER ═════════════════════════════════════════════════════════ */
.footer {{
  padding: 12px 24px 20px;
  font-size: 10px;
  color: var(--ink3);
  border-top: 1px solid var(--border);
  margin-top: 8px;
  display: flex;
  justify-content: space-between;
  align-items: center;
}}
.footer-brand {{ font-weight: 600; color: var(--ink2); letter-spacing: .5px }}

@media print {{
  .fbar, .fbar-actions, .chips {{ display: none !important }}
  body {{ background: #fff }}
  .hdr, thead th {{ -webkit-print-color-adjust: exact; print-color-adjust: exact }}
}}
</style>
</head>
<body>

<!-- HEADER -->
<div class="hdr">
  <div class="hdr-brand">
    <img src="data:image/png;base64,{LOGO_B64}" 
         alt="Casas y Colores" 
         style="height:48px;width:auto;margin-bottom:6px;display:block">
    <div class="hdr-eyebrow" style="margin-top:2px">Contabilidad</div>
    <div class="hdr-title">Comprobantes Pendientes</div>
  </div>
  <div class="hdr-kpis">
    <div class="hdr-kpi">
      <div class="hdr-kpi-lbl">Total SIRE</div>
      <div class="hdr-kpi-val">{total_sire}</div>
      <div class="hdr-kpi-sub">Tipos 01·07·08·30·42·50·53·54</div>
    </div>
    <div class="hdr-kpi">
      <div class="hdr-kpi-lbl">Ya en SAP</div>
      <div class="hdr-kpi-val success">{ya_reg}</div>
      <div class="hdr-kpi-sub">{pct_reg}% registrado</div>
    </div>
    <div class="hdr-kpi">
      <div class="hdr-kpi-lbl">Pendientes</div>
      <div class="hdr-kpi-val danger">{len(pendientes)}</div>
      <div class="hdr-kpi-sub">{pct_pend}% sin ingresar</div>
    </div>
    <div class="hdr-kpi">
      <div class="hdr-kpi-lbl">Total Pendiente</div>
      <div class="hdr-kpi-val danger">S/ {total_cp:,.0f}</div>
      <div class="hdr-kpi-sub">Por ingresar al SAP</div>
    </div>
    <div class="hdr-kpi">
      <div class="hdr-kpi-lbl">Proveedores</div>
      <div class="hdr-kpi-val">{num_prov}</div>
      <div class="hdr-kpi-sub">Distintos con pendientes</div>
    </div>
  </div>
  <div class="hdr-periodo">
    <div class="hdr-mes">{periodo}</div>
    <div class="hdr-fechas">{fecha_min} — {fecha_max}</div>
    <div class="hdr-fechas" style="margin-top:4px">Actualizado: {now_str}</div>
  </div>
</div>

<!-- BARRA PROGRESO -->
<div class="prog-wrap">
  <div class="prog-card">
    <span class="prog-lbl">Avance de registro</span>
    <div class="prog-track">
      <div class="prog-reg"></div>
      <div class="prog-pend"></div>
    </div>
    <div class="prog-legend">
      <div class="prog-legend-item">
        <span class="dot" style="background:var(--green)"></span>
        Registrados en SAP: <strong>{ya_reg}</strong> ({pct_reg}%)
      </div>
      <div class="prog-legend-item">
        <span class="dot" style="background:var(--red)"></span>
        Pendientes: <strong>{len(pendientes)}</strong> ({pct_pend}%)
      </div>
    </div>
  </div>
</div>

<!-- TARJETAS POR TIPO -->
<div class="tipo-cards">{tipo_cards_html}</div>



<!-- TOP PROVEEDORES -->
<div class="sec">
  <div class="sec-hdr">
    <span class="sec-title">Top 10 Proveedores — Mayor cantidad de comprobantes pendientes</span>
    <span class="sec-line"></span>
    <span class="badge red">{num_prov} proveedores</span>
  </div>
  <div class="tbl-wrap">
    <table style="min-width:0;table-layout:fixed;width:100%">
      <thead><tr>
        <th style="width:36%">Proveedor / Razón Social</th>
        <th style="width:18%">Tipos</th>
        <th class="r" style="width:6%">N° CP</th>
        <th class="r tot-h" style="width:20%">Total Pendiente</th>
        <th class="r igv-h" style="width:20%">IGV</th>
      </tr></thead>
      <tbody>{prov_html}</tbody>
    </table>
  </div>
</div>

<!-- DETALLE -->
<div class="sec">
  <div class="sec-hdr">
    <span class="sec-title">Detalle de Comprobantes Pendientes</span>
    <span class="sec-line"></span>
    <span class="badge red" id="badge-count">{len(pendientes)} registros</span>
  </div>

  <div class="fbar">
    <div class="fbar-row">
      <div class="fg fg-wide">
        <label>Razón Social</label>
        <div class="razon-wrap">
          <input type="text" id="f-razon" list="razon-list"
                 placeholder="Escribe o selecciona proveedor…"
                 onkeydown="if(event.key==='Enter')applyFilters()">
          <datalist id="razon-list">{razon_opts}</datalist>
          <button class="btn-buscar" onclick="applyFilters()">
            <svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/></svg>
            Buscar
          </button>
        </div>
      </div>
      <div class="fg fg-wide">
        <label>RUC / Serie / Número</label>
        <input type="text" id="f-texto" placeholder="Ej: 20610024981 · E001 · 1080…" oninput="applyFilters()">
      </div>
      <div class="fbar-actions">
        <button class="btn btn-clear" onclick="clearFilters()">✕ Limpiar</button>
        <button class="btn btn-excel" onclick="exportExcel()">
          <svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="8" y1="13" x2="16" y2="13"/><line x1="8" y1="17" x2="16" y2="17"/></svg>
          Exportar Excel
        </button>
      </div>
    </div>
    <div class="fbar-row">
      <div class="fg">
        <label>Tipo CP</label>
        <select id="f-tipo" onchange="applyFilters()">
          <option value="">Todos los tipos</option>
          {tipo_opts}
        </select>
      </div>
      <div class="fg">
        <label>Fecha Emisión</label>
        <select id="f-fecha" onchange="applyFilters()">
          <option value="">Todas las fechas</option>
          {date_opts}
        </select>
      </div>
      <div class="fg">
        <label>Moneda</label>
        <select id="f-moneda" onchange="applyFilters()">
          <option value="">Todas</option>
          {mon_opts}
        </select>
      </div>
      <div class="fg fg-num">
        <label>Total mínimo</label>
        <input type="number" id="f-min" placeholder="0" step="100" oninput="applyFilters()">
      </div>
      <div class="fg fg-num">
        <label>Total máximo</label>
        <input type="number" id="f-max" placeholder="Sin límite" min="0" step="100" oninput="applyFilters()">
      </div>
    </div>
    <div class="chips" id="chips"></div>
  </div>

  <div class="res-bar">
    <div class="res-count">Mostrando <strong id="rc-n">{len(pendientes)}</strong> de {len(pendientes)} comprobantes pendientes</div>
    <div class="res-totals">
      <div class="res-item"><span class="res-lbl">Base Imponible</span><span class="res-val" id="s-bi">S/ {total_bi:,.2f}</span></div>
      <div class="res-item"><span class="res-lbl">IGV</span><span class="res-val igv-v" id="s-igv">S/ {total_igv:,.2f}</span></div>
      <div class="res-item"><span class="res-lbl">Total Pendiente</span><span class="res-val tot-v" id="s-tot">S/ {total_cp:,.2f}</span></div>
    </div>
  </div>

  <div class="tbl-wrap">
    <table>
      <thead><tr>
        <th class="col-fecha">Fecha</th>
        <th class="col-tipo">Tipo CP</th>
        <th class="col-serie">Serie</th>
        <th class="col-ncp">N° CP</th>
        <th class="col-ruc">RUC</th>
        <th>Proveedor / Razón Social</th>
        <th class="r">Base Imponible</th>
        <th class="r igv-h">IGV / IPM</th>
        <th class="r tot-h col-total">Total CP</th>
        <th class="col-mon">Mon.</th>
      </tr></thead>
      <tbody id="tbody-fact"></tbody>
    </table>
  </div>
</div>

<div class="footer">
  <span class="footer-brand">SIRE ↔ SAP · Casas y Colores · Contabilidad</span>
  <span>{periodo} · {len(pendientes)} pendientes de {total_sire} · Tipos 01·07·08·30·42·50·53·54 · Uso interno</span>
  <span>Actualizado: {now_str}</span>
</div>

<script src="https://cdn.jsdelivr.net/npm/xlsx-js-style@1.2.0/dist/xlsx.bundle.js"></script>
<script>
const DATA={json.dumps(js_data,ensure_ascii=False)};
const TOTAL=DATA.length;
const TIPO_LABEL={json.dumps({t:tipo_label(t) for t in ['01','07','08','30','42','50','53','54']},ensure_ascii=False)};
const TIPO_COLOR={json.dumps(TIPO_COLOR,ensure_ascii=False)};
const TIPO_BG={json.dumps(TIPO_BG,ensure_ascii=False)};
const fmt=v=>v.toLocaleString('es-PE',{{minimumFractionDigits:2,maximumFractionDigits:2}});

function tipoBadge(t){{
  const col=TIPO_COLOR[t]||'#374151';
  const bg=TIPO_BG[t]||'#F2F4F7';
  const lbl=TIPO_LABEL[t]||t;
  return`<span class="tipo-badge" style="background:${{bg}};color:${{col}};border-color:${{col}}44">${{lbl}}</span>`;
}}

function renderRows(list){{
  const tb=document.getElementById('tbody-fact');
  if(!list.length){{
    tb.innerHTML='<tr><td colspan="10" class="empty-td">Sin comprobantes pendientes con los filtros aplicados.</td></tr>';
    return;
  }}
  tb.innerHTML=list.map(d=>{{
    const prov=d.proveedor.length>55?d.proveedor.slice(0,55)+'…':d.proveedor;
    return`<tr>
      <td class="mono col-fecha">${{d.fecha}}</td>
      <td class="col-tipo">${{tipoBadge(d.tipo)}}</td>
      <td class="mono col-serie">${{d.serie}}</td>
      <td class="mono col-ncp">${{d.nro}}</td>
      <td class="mono col-ruc">${{d.ruc}}</td>
      <td class="proveedor-cell" title="${{d.proveedor}}">${{prov}}</td>
      <td class="num col-importe">${{fmt(d.bi)}}</td>
      <td class="num igv col-importe">${{fmt(d.igv)}}</td>
      <td class="num total-cp col-importe" style="${{d.total<0?'color:#7B2020;background:#FDF3F2':''}}">
        ${{d.total<0?'- ':''}}${{fmt(Math.abs(d.total))}}</td>
      <td class="mono col-mon" style="font-size:12px;font-weight:600">${{d.moneda}}</td>
    </tr>`;
  }}).join('');
}}

function updateSummary(list){{
  document.getElementById('rc-n').textContent=list.length;
  document.getElementById('badge-count').textContent=list.length+' registros';
  document.getElementById('s-bi').textContent=fmt(list.reduce((a,d)=>a+d.bi,0));
  document.getElementById('s-igv').textContent=fmt(list.reduce((a,d)=>a+d.igv,0));
  document.getElementById('s-tot').textContent=fmt(list.reduce((a,d)=>a+d.total,0));
}}

function getActive(){{
  return{{
    razon:  document.getElementById('f-razon').value.trim(),
    texto:  document.getElementById('f-texto').value.trim().toLowerCase(),
    tipo:   document.getElementById('f-tipo').value,
    fecha:  document.getElementById('f-fecha').value,
    moneda: document.getElementById('f-moneda').value,
    min:    document.getElementById('f-min').value !== '' ? parseFloat(document.getElementById('f-min').value) : -Infinity,
    max:    parseFloat(document.getElementById('f-max').value)||Infinity,
  }};
}}

function getFiltered(){{
  const f=getActive();
  return DATA.filter(d=>{{
    if(f.razon  && !d.proveedor.toLowerCase().includes(f.razon.toLowerCase())) return false;
    if(f.texto  && !d.ruc.includes(f.texto) && !d.proveedor.toLowerCase().includes(f.texto)
                && !d.serie.toLowerCase().includes(f.texto) && !d.nro.includes(f.texto)) return false;
    if(f.tipo   && d.tipo   !== f.tipo)   return false;
    if(f.fecha  && d.fecha  !== f.fecha)  return false;
    if(f.moneda && d.moneda !== f.moneda) return false;
    const absTotal = Math.abs(d.total);
    if(f.min !== -Infinity && absTotal < f.min) return false;
    if(f.max !== Infinity  && absTotal > f.max) return false;
    return true;
  }});
}}

function renderChips(){{
  const f=getActive(); const chips=[];
  const clr=id=>()=>{{document.getElementById(id).value='';applyFilters();}};
  if(f.razon)  chips.push(['Razón: '+f.razon.slice(0,25), clr('f-razon')]);
  if(f.texto)  chips.push(['Texto: '+f.texto.slice(0,22), clr('f-texto')]);
  if(f.tipo)   chips.push(['Tipo: '+(TIPO_LABEL[f.tipo]||f.tipo), clr('f-tipo')]);
  if(f.fecha)  chips.push(['Fecha: '+f.fecha, clr('f-fecha')]);
  if(f.moneda) chips.push(['Moneda: '+f.moneda, clr('f-moneda')]);
  if(f.min>0)  chips.push(['Mín: S/'+f.min, clr('f-min')]);
  if(f.max<Infinity) chips.push(['Máx: S/'+f.max, clr('f-max')]);
  const el=document.getElementById('chips');
  el.innerHTML=chips.map(([l],i)=>`<div class="chip">${{l}}<span class="chip-x" data-i="${{i}}">✕</span></div>`).join('');
  el.querySelectorAll('.chip-x').forEach(x=>{{x.onclick=chips[+x.dataset.i][1];}});
}}

function applyFilters(){{
  const l=getFiltered(); renderRows(l); updateSummary(l); renderChips();
}}

function clearFilters(){{
  ['f-razon','f-texto'].forEach(id=>document.getElementById(id).value='');
  ['f-tipo','f-fecha','f-moneda'].forEach(id=>document.getElementById(id).value='');
  ['f-min','f-max'].forEach(id=>document.getElementById(id).value='');
  applyFilters();
}}

function exportExcel(){{
  const list = getFiltered();
  const WB   = XLSX.utils.book_new();
  const hoy  = new Date().toLocaleDateString('es-PE');
  const ISO  = new Date().toISOString().slice(0,10);

  // ── Helpers de estilo ─────────────────────────────────────────────
  const NAVY  = '1E3A5F'; const WHITE = 'FFFFFF'; const LGRAY = 'F2F4F7';
  const RED   = '7B2020'; const GREEN = '1A5C35'; const IGVBL = '1E4878';
  const TOTBG = '6B2020'; const BORD  = 'C8CDD8';

  const border = {{
    top:   {{style:'thin', color:{{rgb:BORD}}}},
    bottom:{{style:'thin', color:{{rgb:BORD}}}},
    left:  {{style:'thin', color:{{rgb:BORD}}}},
    right: {{style:'thin', color:{{rgb:BORD}}}},
  }};

  function cell(v, s) {{ return {{v, s, t: typeof v==='number'?'n':'s'}}; }}

  const NF = '#,##0.00';

  // ── Construir filas ───────────────────────────────────────────────
  // Fila 1: título
  const R1 = [
    cell('COMPROBANTES PENDIENTES DE INGRESO AL SAP — {periodo}', {{
      font:{{bold:true, sz:13, color:{{rgb:WHITE}}, name:'Arial'}},
      fill:{{fgColor:{{rgb:NAVY}}, patternType:'solid'}},
      alignment:{{horizontal:'left', vertical:'center'}},
    }}),
    ...Array(13).fill(cell('',{{fill:{{fgColor:{{rgb:NAVY}},patternType:'solid'}}}}))
  ];

  // Fila 2: subtítulo
  const R2 = [
    cell('Casas y Colores  ·  Cruce SIRE vs SAP  ·  Tipos: 01·07·08·30·42·50·54  ·  ' + hoy, {{
      font:{{italic:true, sz:9, color:{{rgb:WHITE}}, name:'Arial'}},
      fill:{{fgColor:{{rgb:'254878'}}, patternType:'solid'}},
      alignment:{{horizontal:'left', vertical:'center'}},
    }}),
    ...Array(13).fill(cell('',{{fill:{{fgColor:{{rgb:'254878'}},patternType:'solid'}}}}))
  ];

  // Fila 3: vacía
  const R3 = Array(14).fill(cell('',{{}}));

  // Fila 4: cabeceras
  const hdrs = ['FECHA','TIPO CP','SERIE','N° CP','RUC','PROVEEDOR / RAZÓN SOCIAL','BASE IMPONIBLE','IGV / IPM','TOTAL CP','MONEDA','ORIG. TIPO','ORIG. FECHA','ORIG. SERIE','ORIG. N°'];
  const hdrBg = [NAVY,NAVY,NAVY,NAVY,NAVY,NAVY,NAVY,IGVBL,TOTBG,NAVY,'4A2060','4A2060','4A2060','4A2060'];
  const R4 = hdrs.map((h,i) => cell(h, {{
    font:    {{bold:true, sz:9, color:{{rgb:WHITE}}, name:'Arial'}},
    fill:    {{fgColor:{{rgb:hdrBg[i]}}, patternType:'solid'}},
    alignment:{{horizontal: i>=6&&i<=8?'right':'left', vertical:'center'}},
    border,
  }}));

  // Filas de datos
  const dataRows = list.map((d, idx) => {{
    const isEven = idx % 2 === 0;
    const rowBg  = isEven ? WHITE : LGRAY;
    const numStyle = (bg) => ({{
      font:{{sz:9, name:'Arial', color:{{rgb:RED}}, bold:true}},
      fill:{{fgColor:{{rgb:bg}}, patternType:'solid'}},
      alignment:{{horizontal:'right', vertical:'center'}},
      border, numFmt: NF,
    }});
    const txtStyle = (bg, clr='111827', bold=false) => ({{
      font:{{sz:9, name:'Arial', color:{{rgb:clr}}, bold}},
      fill:{{fgColor:{{rgb:bg}}, patternType:'solid'}},
      alignment:{{horizontal:'left', vertical:'center'}},
      border,
    }});
    return [
      cell(d.fecha,     txtStyle(rowBg)),
      cell(TIPO_LABEL[d.tipo]||d.tipo, txtStyle(rowBg)),
      cell(d.serie,     txtStyle(rowBg)),
      cell(d.nro,       txtStyle(rowBg)),
      cell(d.ruc,       txtStyle(rowBg, '6B7280')),
      cell(d.proveedor, txtStyle(rowBg)),
      {{v:d.bi,   t:'n', s:{{font:{{sz:9,name:'Arial',color:{{rgb:'111827'}}}},fill:{{fgColor:{{rgb:'F0F5FF'}},patternType:'solid'}},alignment:{{horizontal:'right',vertical:'center'}},border,numFmt:NF}}}},
      {{v:d.igv,  t:'n', s:{{font:{{sz:9,name:'Arial',color:{{rgb:'111827'}}}},fill:{{fgColor:{{rgb:'EEF2FA'}},patternType:'solid'}},alignment:{{horizontal:'right',vertical:'center'}},border,numFmt:NF}}}},
      {{v:d.total,t:'n', s:numStyle('FDF3F2')}},
      cell(d.moneda, txtStyle(rowBg)),
      cell(d.orig_tipo  ||'', {{font:{{sz:9,name:'Arial',color:{{rgb:d.orig_tipo?'4A2060':'9CA3AF'}}}},fill:{{fgColor:{{rgb:d.orig_tipo?'F3EEFF':rowBg}},patternType:'solid'}},alignment:{{horizontal:'left',vertical:'center'}},border}}),
      cell(d.orig_fecha ||'', {{font:{{sz:9,name:'Arial',color:{{rgb:d.orig_tipo?'4A2060':'9CA3AF'}}}},fill:{{fgColor:{{rgb:d.orig_tipo?'F3EEFF':rowBg}},patternType:'solid'}},alignment:{{horizontal:'left',vertical:'center'}},border}}),
      cell(d.orig_serie ||'', {{font:{{sz:9,name:'Arial',color:{{rgb:d.orig_tipo?'4A2060':'9CA3AF'}}}},fill:{{fgColor:{{rgb:d.orig_tipo?'F3EEFF':rowBg}},patternType:'solid'}},alignment:{{horizontal:'left',vertical:'center'}},border}}),
      cell(d.orig_nro   ||'', {{font:{{sz:9,name:'Arial',color:{{rgb:d.orig_tipo?'4A2060':'9CA3AF'}}}},fill:{{fgColor:{{rgb:d.orig_tipo?'F3EEFF':rowBg}},patternType:'solid'}},alignment:{{horizontal:'left',vertical:'center'}},border}}),
    ];
  }});

  // Fila vacía
  const RSEP = Array(14).fill(cell('',{{}}));

  // Fila totales
  const totBI  = list.reduce((a,d)=>a+d.bi,0);
  const totIGV = list.reduce((a,d)=>a+d.igv,0);
  const totCP  = list.reduce((a,d)=>a+d.total,0);
  const totStyle = (right=false, numFmt=null) => {{
    const s = {{
      font:{{bold:true, sz:10, color:{{rgb:WHITE}}, name:'Arial'}},
      fill:{{fgColor:{{rgb:GREEN}}, patternType:'solid'}},
      alignment:{{horizontal: right?'right':'left', vertical:'center'}},
      border,
    }};
    if (numFmt) s.numFmt = numFmt;
    return s;
  }};
  const RTOT = [
    cell('TOTALES',                 totStyle()),
    cell('',                        totStyle()),
    cell('',                        totStyle()),
    cell('',                        totStyle()),
    cell(list.length+' comprobantes', totStyle(false)),
    cell('',                        totStyle()),
    {{v:totBI,  t:'n', s:totStyle(true, NF)}},
    {{v:totIGV, t:'n', s:totStyle(true, NF)}},
    {{v:totCP,  t:'n', s:totStyle(true, NF)}},
    cell('',                        totStyle()),
  ];

  // ── Armar sheet ───────────────────────────────────────────────────
  const allRows = [R1, R2, R3, R4, ...dataRows, RSEP, RTOT];
  const WS = XLSX.utils.aoa_to_sheet(allRows);

  // Anchos
  WS['!cols'] = [
    {{wch:12}},{{wch:20}},{{wch:8}},{{wch:12}},{{wch:14}},
    {{wch:50}},{{wch:16}},{{wch:14}},{{wch:14}},{{wch:8}},
    {{wch:10}},{{wch:14}},{{wch:10}},{{wch:10}},
  ];

  // Merges título y subtítulo
  WS['!merges'] = [
    {{s:{{r:0,c:0}}, e:{{r:0,c:13}}}},
    {{s:{{r:1,c:0}}, e:{{r:1,c:13}}}},
  ];

  // Alturas
  WS['!rows'] = [
    {{hpt:22}}, // R1 título
    {{hpt:14}}, // R2 subtítulo
    {{hpt:6}},  // R3 vacía
    {{hpt:16}}, // R4 cabeceras
    ...list.map(()=>( {{hpt:14}} )),
    {{hpt:6}},  // sep
    {{hpt:16}}, // totales
  ];

  XLSX.utils.book_append_sheet(WB, WS, 'Pendientes SAP');
  XLSX.writeFile(WB, 'Pendientes_SAP_' + ISO + '.xlsx');
}}

applyFilters();
</script>
</body>
</html>"""


OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'index.html')
with open(OUT, 'w', encoding='utf-8') as f:
    f.write(HTML)
print(f"\n{'='*50}\n  ✅ index.html listo ({len(HTML)//1024} KB)\n  📅 {periodo} · ❌ {len(pendientes)} pendientes · 💰 S/ {total_cp:,.2f}\n{'='*50}")
